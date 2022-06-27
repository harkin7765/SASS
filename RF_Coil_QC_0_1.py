"""
Code for calculating the SNR and Uniformity of multi element coils. 
For more details on the rationalle behind this analysis please see the documents
provided by Moriel NessAiver from SimplyPhysics: 
-"Compensating for Integer Truncation When Measuring the SNR of Low Noise Magnetic Resonance Images"
-"RF Coil Testing - The Simply Physics way"

Both individual element and combined element images should be supplied to the code.
Combined elemnt inmages should be produced via a sum of squares method  
Images should be acquired with all post-processing turned off.If this is not 
possible the image processing should be the same as used clinically.

Images are sorted into individual element images and combined images.  For 
both sets the SNR is calculated and for the combined images the uniformity is also calculated.

The user is asked to confirm if the automatic mask produced is acceptable.  
This mask is generated using the combined element image. and is used on all 
images so all the DICOMs selected should be acquired with the phantom in the same location using the same sequence parameters  
The thresholds should be adjusted until the displayed mask covers the entirity of the phantom.
To produce the phantom mask the accepted mask is contracted by 4 pixels and to 
produce the in air mask the accepted mask is inverted and contracted by 5 pixels.

SNR is calculated via 3 methods: the NEMA subtraction method which requires 
two identically acquired images and two methods which only require a sinle image. 
The two single image methods differ in the way the noise is approximated: one 
assumes the noise is roportional to the average signal in air and one assumes 
the noise is proportional to the standard deviation of the signal in air.

Depending on the number of elements used to generate image the SNR is scaled by different factors derived from
"Constantinides, C. D., Atalar, E., McVeigh, E. R. (1997). Signal-to-noise measurements
from magnitude images in NMR phased-arrays. Magnetic Resonance in Medicine, 38, 852-
857."

Uniformity is calculated via the method outlined in IPEM Report 112 and is refereed
to as the integral uniformity method. A low pass filter is applied to the image and the "phantom signal
mask" (used for SNR calculation) is applied to the image so only voxels only within the phantom are
considered. 

If the user requires the results to be exported to excel. 12 Excel files are 
created/ modified.  For each type of SNR calculation there are 4 excel files. 
One for the individual element data, one for the combined element data, one for 
an average of the individual element data and one for an average of the combined 
element data.

If the user wants PNG files to be produced for the purpose of reporting 3 sets are produced:
    -PixelArray images: Images based on DICOM.PixelArray for each combined and individual element image
    -Tables: Comparing current results to previously established baselines
    -Graphs: Graph of historic and current data
Tables and Graphs require baseline values for each parameter to be established in baseline spreadsheets.
"""

import pydicom
from pydicom.tag import Tag
import os
import shutil
import tkinter as tk
import numpy as np
from matplotlib import pyplot as plt
from PIL import ImageTk,Image
from tkinter import filedialog, messagebox
import skimage.segmentation as seg
import openpyxl
import cv2
import pandas as pd
import dataframe_image as dfi             
                

root = tk.Tk()
root.title("Moriel QC")
root.geometry("800x800")

class main:
    #root path to search for images
    Quarterly_path = "S:/Radiology-Directorate/IRIS/Non-Ionising/MRI/QA and Calibration/Quarterly QC/SNR QC/Images/PHT_Ambition_X/"
    class initialise_analysis:    
        
        #Acceptable list of sequence names
        SNR_sequence_names = ["DelRec - 256 AP"]# "SE_SNR_ND", "SE_SNR_SAG_ND", "SE_SNR_ND", "SE_SNR", "t1_se_tra", "DelRec - 256 AP", "DelRec - SE_SNR"]
        #root location to store PNG images e.g. graphs, tables and magnitude images
        png_archive = "S:/Radiology-Directorate/IRIS/Non-Ionising/MRI/QA and Calibration/Quarterly QC/SNR QC/PNG_images/"
        scanner_ID_dict = {"45504":"PHT_ElitionX_MR3", "47009": "PHT_Ambition_X", "169672":"PHT_Avanto","169696":"DCH_Avanto_MR1", "183109": "DCH_Sola_MR2", "202541": "RBH_Sola_MR2", "183188": "RBH_Sola_MR3", "64114": "RBH_Essenza_MR3", "169641": "RBH_Avanto_MR1", "41113": "RBH_Aera_MR2", "42044": "PHT_Ingenia3T_MR3", "2016": "AECC_Paramed_MROPEN"}        
        coil_dict = {
            "45504": {
                 "Anterior_74":{"n_slices":1, "n_elements":28, "lower_threshold":0.1, "upper_threshold":0.1},
                 "Anterior_79":{"n_slices":1, "n_elements":28, "lower_threshold":0.1, "upper_threshold":0.1}, 
                 #"Breast":{"n_slices":1, "n_elements":15, "lower_threshold":0.01, "upper_threshold":0.1},
                 "Breast_53":{"n_slices":1, "n_elements":16, "lower_threshold":0.01, "upper_threshold":0.1},
                 "Extremity":{"n_slices":1, "n_elements":8, "lower_threshold":0.01, "upper_threshold":0.05},
                 "Flex-M_234":{"n_slices":1, "n_elements":1, "lower_threshold":0.1, "upper_threshold":0.1},
                 "Flex-M_341":{"n_slices":1, "n_elements":1, "lower_threshold":0.1, "upper_threshold":0.1},
                 "Foot":{"n_slices":1, "n_elements":16, "lower_threshold":0.01, "upper_threshold":0.05},
                 "Head_Neck_42_Cor":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.1},
                 "Head_Neck_42_Sag":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.1}, 
                 "Head_Neck_42_Tra":{"n_slices":1, "n_elements":15, "lower_threshold":0.1, "upper_threshold":0.1},
                 #"Head_Neck_42_Tra":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.1},
                 "Knee_2271":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.1},
                 "Shoulder_89":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.06}, 
                 "Wrist":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.02},
                 "Head":{"n_slices":1, "n_elements":14, "lower_threshold":0.1, "upper_threshold":0.1},
                 },            
            "47009": {
                 "Anterior171":{"n_slices":1, "n_elements":28, "lower_threshold":0.1, "upper_threshold":0.1},
                 "Anterior603":{"n_slices":1, "n_elements":28, "lower_threshold":0.1, "upper_threshold":0.1}, 
                 "Extremity_430":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.1},
                 "Foot_Ankle":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.1},
                 "Head":{"n_slices":1, "n_elements":15, "lower_threshold":0.1, "upper_threshold":0.1},#"n_elements":16
                 "Head_Neck_Cor":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.1}, 
                 "Head_Neck_Sag":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.1}, 
                 "Head_Neck_Tra":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.25}, 
                 "Knee_698":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.1}, 
                 "Shoulder_126":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.06}, 
                 "Spine":{"n_slices":1, "n_elements":12, "lower_threshold":0.1, "upper_threshold":0.2}#"upper_threshold":0.1
                 },
            "169672": {
                 "Body18_33362":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.075}, 
                 "Body18_Long_31201":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.075},
                 "Bore_Body_Cor":{"n_slices":1, "n_elements":2, "lower_threshold":0.1, "upper_threshold":0.25},
                 "Bore_Body_Sag":{"n_slices":1, "n_elements":2, "lower_threshold":0.1, "upper_threshold":0.25},
                 "Bore_Body_Tra":{"n_slices":1, "n_elements":2, "lower_threshold":0.1, "upper_threshold":0.25},
                 "Breast":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.045},
                 "Flex_Large":{"n_slices":1, "n_elements":4, "lower_threshold":0.1, "upper_threshold":0.1},
                 "Flex_Small":{"n_slices":1, "n_elements":4, "lower_threshold":0.1, "upper_threshold":0.1},
                 "Foot_Ankle":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.16}, 
                 "Hand_Wrist":{"n_slices":1, "n_elements":16, "lower_threshold":0.01, "upper_threshold":0.03}, 
                 "Head_Neck_Cor":{"n_slices":1, "n_elements":20, "lower_threshold":0.1, "upper_threshold":0.07},
                 #"Head_Neck_Sag":{"n_slices":1, "n_elements":20, "lower_threshold":0.1, "upper_threshold":0.05}, 
                 #"Head_Neck_Tra":{"n_slices":1, "n_elements":20, "lower_threshold":0.1, "upper_threshold":0.25},
                 "Knee":{"n_slices":1, "n_elements":15, "lower_threshold":0.1, "upper_threshold":0.1},
                 "PA_1_3":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.05}, 
                 "PA_4_6":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.05},
                 "Shoulder_Large":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.1},
                 "Shoulder_Small":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.1},
                 "Spine_1_4":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.25},
                 "Spine_5_8":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.25}, 
                 },
            "169696": {
                 "Body_18":{"n_slices":1, "n_elements":18, "lower_threshold":0.015, "upper_threshold":0.055}, 
                 "Body18_Long":{"n_slices":1, "n_elements":18, "lower_threshold":0.017, "upper_threshold":0.075},
                 "Body30":{"n_slices":1, "n_elements":30, "lower_threshold":0.1, "upper_threshold":0.075},
                 "Bore_Body":{"n_slices":1, "n_elements":2, "lower_threshold":0.1, "upper_threshold":0.25},
                 "Breast":{"n_slices":1, "n_elements":4, "lower_threshold":0.1, "upper_threshold":0.045},
                 "Flex_Large":{"n_slices":1, "n_elements":4, "lower_threshold":0.1, "upper_threshold":0.1},
                 "Flex_Small":{"n_slices":1, "n_elements":4, "lower_threshold":0.1, "upper_threshold":0.1},
                 "Foot_Ankle":{"n_slices":1, "n_elements":16, "lower_threshold":0.01, "upper_threshold":0.14}, 
                 "Hand_Wrist":{"n_slices":1, "n_elements":16, "lower_threshold":0.01, "upper_threshold":0.03}, 
                 "Head_Neck_Cor":{"n_slices":1, "n_elements":20, "lower_threshold":0.1, "upper_threshold":0.07},
                 "Head_Neck_Sag":{"n_slices":1, "n_elements":20, "lower_threshold":0.1, "upper_threshold":0.05}, 
                 "Head_Neck_Tra":{"n_slices":1, "n_elements":20, "lower_threshold":0.1, "upper_threshold":0.25},
                 "Knee":{"n_slices":1, "n_elements":15, "lower_threshold":0.1, "upper_threshold":0.1},
                 "PA_1_3":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.05}, 
                 "PA_4_6":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.05},
                 "Shoulder_Large":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.1},
                 "Shoulder_Small":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.1},
                 "Spine_1_4":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.25},
                 "Spine_5_8":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.25}, 
                 },
            "183109":{
                      "Body_31044":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.1},
                      "Body_34112":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.1},
                      "Bore_Body":{"n_slices":1, "n_elements":2, "lower_threshold":0.1, "upper_threshold":0.25},
                      "Breast":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.06},
                      "Flex_Large":{"n_slices":1, "n_elements":4, "lower_threshold":0.1, "upper_threshold":0.1},
                      "Flex_Small":{"n_slices":1, "n_elements":4, "lower_threshold":0.1, "upper_threshold":0.1},
                      "Foot_Ankle":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.1},
                      "Head_Neck":{"n_slices":1, "n_elements":20, "lower_threshold":0.1, "upper_threshold":0.055},
                      "Knee":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.1},
                      "PA_1-3":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.04},
                      "PA_4-6":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.04},
                      "Shoulder":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.1},
                      "Spine_1-4":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.25},
                      "Spine_5-8":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.25},
                      "Wrist":{"n_slices":1, "n_elements":16, "lower_threshold":0.01, "upper_threshold":0.04}},
            "183188":{
                    "Body_31117":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.1},
                    "Body_34437":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.1},
                    "Bore_Body_Cor":{"n_slices":1, "n_elements":2, "lower_threshold":0.1, "upper_threshold":0.25},
                    "Bore_Body_Sag":{"n_slices":1, "n_elements":2, "lower_threshold":0.1, "upper_threshold":0.25},
                    "Bore_Body_Tra":{"n_slices":1, "n_elements":2, "lower_threshold":0.1, "upper_threshold":0.25},
                    "Flex_Large":{"n_slices":1, "n_elements":4, "lower_threshold":0.1, "upper_threshold":0.1},
                    "Flex_Small":{"n_slices":1, "n_elements":4, "lower_threshold":0.1, "upper_threshold":0.1},
                    "Foot_Ankle":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.1},
                    "Head_Neck_COR":{"n_slices":1, "n_elements":20, "lower_threshold":0.1, "upper_threshold":0.055},
                    "Head_Neck_SAG":{"n_slices":1, "n_elements":20, "lower_threshold":0.1, "upper_threshold":0.05},
                    "Head_Neck_TRA":{"n_slices":1, "n_elements":20, "lower_threshold":0.1, "upper_threshold":0.25},
                    "Knee":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.1},
                    "PA_1-3":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.05},
                    "PA_4-6":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.05},
                    "Shoulder":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.1},
                    "Spine_1-4":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.25},
                    "Spine_5-8":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.25},
                    "Wrist":{"n_slices":1, "n_elements":16, "n_elements":16, "lower_threshold":0.01, "upper_threshold":0.04}},
            "202541":{
                    "Body_18_34823":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.1},
                    "Body_18_Long_31212":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.1},
                    "Bore_Body_Cor":{"n_slices":1, "n_elements":2, "lower_threshold":0.1, "upper_threshold":0.25},
                    "Bore_Body_Sag":{"n_slices":1, "n_elements":2, "lower_threshold":0.1, "upper_threshold":0.25},
                    "Bore_Body_Tra":{"n_slices":1, "n_elements":2, "lower_threshold":0.1, "upper_threshold":0.25},
                    "Flex_Large":{"n_slices":1, "n_elements":4, "lower_threshold":0.1, "upper_threshold":0.1},
                    "Flex_Small":{"n_slices":1, "n_elements":4, "lower_threshold":0.1, "upper_threshold":0.1},
                    "Foot_Ankle":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.1},
                    "Hand_Wrist":{"n_slices":1, "n_elements":16, "n_elements":16, "lower_threshold":0.01, "upper_threshold":0.04},
                    "Head_Neck":{"n_slices":1, "n_elements":20, "lower_threshold":0.1, "upper_threshold":0.055},
                    "Knee":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.1},
                    "PA_1-3":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.05},
                    "PA_4-6":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.05},
                    "Shoulder":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.1},
                    "Spine_1-4":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.2},
                    "Spine_5-8":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.2}},
             "169641": {
                 "Body18_3198":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.075}, 
                 "Body18_Long_30223":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.075},
                 "Bore_Body_COR":{"n_slices":1, "n_elements":2, "lower_threshold":0.1, "upper_threshold":0.25},
                 "Bore_Body_SAG":{"n_slices":1, "n_elements":2, "lower_threshold":0.1, "upper_threshold":0.25},
                 "Bore_Body_TRA":{"n_slices":1, "n_elements":2, "lower_threshold":0.1, "upper_threshold":0.25},
                 "Flex_Large":{"n_slices":1, "n_elements":4, "lower_threshold":0.1, "upper_threshold":0.1},
                 "Flex_Small":{"n_slices":1, "n_elements":4, "lower_threshold":0.1, "upper_threshold":0.1},
                 "Foot_Ankle":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.1}, 
                 "Hand_Wrist":{"n_slices":1, "n_elements":16, "lower_threshold":0.01, "upper_threshold":0.04}, 
                 "Head_Neck_Cor":{"n_slices":1, "n_elements":20, "lower_threshold":0.1, "upper_threshold":0.06},
                 "Head_Neck_Sag":{"n_slices":1, "n_elements":20, "lower_threshold":0.1, "upper_threshold":0.05}, 
                 "Head_Neck_Tra":{"n_slices":1, "n_elements":20, "lower_threshold":0.1, "upper_threshold":0.25},
                 "Knee":{"n_slices":1, "n_elements":15, "lower_threshold":0.1, "upper_threshold":0.1},
                 "PA_1_3":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.05}, 
                 "PA_4_6":{"n_slices":1, "n_elements":18, "lower_threshold":0.1, "upper_threshold":0.05},
                 "Shoulder_Large":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.1},
                 "Shoulder_Small":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.1},
                 "Spine_1_4":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.25},
                 "Spine_5_8":{"n_slices":1, "n_elements":16, "lower_threshold":0.1, "upper_threshold":0.25},                 
                 }, 
             "41113": {
                 "Neck":{"n_slices":1, "n_elements":1}, 
                 "Shoulder_Large":{"n_slices":1, "n_elements":16}, 
                 "Shoulder_Small":{"n_slices":1, "n_elements":16},
                 "Foot_Ankle":{"n_slices":2, "n_elements":16}, 
                 "Hand_Wrist":{"n_slices":3, "n_elements":16}, 
                 "Knee":{"n_slices":1, "n_elements":15}, 
                 "Spine_1_4":{"n_slices":3, "n_elements":16},
                 "Spine_5_8":{"n_slices":3, "n_elements":16}, 
                 "PA_1_3":{"n_slices":3, "n_elements":18}, 
                 "PA_4_6":{"n_slices":3, "n_elements":18}, 
                 "Flex_Large":{"n_slices":1, "n_elements":4}, 
                 "Flex_Small":{"n_slices":1, "n_elements":4}, 
                 "Body18_30058":{"n_slices":3, "n_elements":18}, 
                 "Body18_32681":{"n_slices":3, "n_elements":18}
                 },
             "42044": {
                 "Anterior0232":{"n_slices":4, "n_elements":27}, 
                 "Anterior6039":{"n_slices":4, "n_elements":27},  
                 "Foot_Ankle":{"n_slices":2, "n_elements":8}, 
                 "Head_Neck":{"n_slices":1, "n_elements":15}, 
                 "Knee":{"n_slices":1, "n_elements":16}, 
                 "Posterior":{"n_slices":4, "n_elements":12}, 
                 "Shoulder":{"n_slices":1, "n_elements":1}, 
                 "Small_Extremeties":{"n_slices":1, "n_elements":16}, 
                 "Wrist":{"n_slices":1, "n_elements":8}, 
                 },
             "64114": {
                 "Body_MATRIX":{"n_slices":1, "n_elements":1}, 
                 "Extremity":{"n_slices":1, "n_elements":1}, 
                 "Flex_Large":{"n_slices":1, "n_elements":1}, 
                 "Flex_Small":{"n_slices":1, "n_elements":1}, 
                 "Foot_Ankle":{"n_slices":1, "n_elements":1}, 
                 "Head":{"n_slices":1, "n_elements":1}, 
                 "Neck":{"n_slices":1, "n_elements":1}, 
                 "Shoulder_Large":{"n_slices":1, "n_elements":1}, 
                 "Shoulder_Small":{"n_slices":1, "n_elements":1}, 
                 "Spine":{"n_slices":1, "n_elements":1}, 
                 "Wrist":{"n_slices":1, "n_elements":1}
                 },

             "2016":{ 
                 "C-SPINE":{"n_slices":1, "n_elements":1}, 
                 "FLAT":{"n_slices":1, "n_elements":1}, 
                 "KNEE":{"n_slices":1, "n_elements":1}, 
                 "MP-LOOP":{"n_slices":1, "n_elements":1}, 
                 "SPINE":{"n_slices":1, "n_elements":1}, 
                 "WRIST":{"n_slices":1, "n_elements":1}
                 }
             }
            
        def __init__(self, images_root_dir):
                """
                Parameters
                ----------
                images_root_dir : str
                    Directory to open file dialogue for selecting directory 
                    containing all dicom files.
                """
                self.base_directory = self.Select_Directory(images_root_dir, 
                                                            "Select directory containing all dicom images to be analysed.")
                #Ask if the user wants to copy the images selected to the archive directory in a sorted format
                self.Ask_Archive()
                if self.archive == True:
                    self.archive_directory = self.Select_Directory(os.path.split(self.base_directory)[0], 
                                                          "Select archive directory.  This directory should contain all previous archived images.")
                #Ask if user wants to produce PNG images.  This includes magnitude images
                #tables of results and graphs tracking SNR and uniformity progress.
                #tables and graphs require baseline spreadsheets to exist
                self.Ask_Figures()
                
                #Ask if results should be exported to excel
                self.Export_To_Excel()
                #get all dicom files inside base_directory
                self.dcm_paths = self.Get_Files(self.base_directory)
                self.dcm_dict = self.Get_Dicom_Dict(self.dcm_paths)
                self.Get_Coil_Name(self.dcm_dict)
                #Sort Dicoms (the sort method varies depending on the scanner)
                if self.scanner_ID == "183188":
                    self.sorted_dcm_dict = self.Sort_Dicom_Dict_Siemens_Sola_RBH(self.dcm_dict)
                if self.scanner_ID == "202541":
                    self.sorted_dcm_dict = self.Sort_Dicom_Dict_Siemens_Sola_DCH(self.dcm_dict)
                elif self.scanner_ID == "183109":
                    self.sorted_dcm_dict = self.Sort_Dicom_Dict_Siemens_Sola_DCH(self.dcm_dict)
                elif self.scanner_ID == "169641":
                    self.sorted_dcm_dict = self.Sort_Dicom_Dict_Siemens_Avanto_RBH_PHT(self.dcm_dict)
                elif self.scanner_ID == "169672":
                    self.sorted_dcm_dict = self.Sort_Dicom_Dict_Siemens_Avanto_RBH_PHT(self.dcm_dict)
                elif  self.scanner_ID == "169696":
                    self.sorted_dcm_dict = self.Sort_Dicom_Dict_Siemens_Avanto_DCH(self.dcm_dict)
                elif self.scanner_ID == "47009":
                    self.sorted_dcm_dict = self.Sort_Dicom_Dict_Ambition_X_PHT(self.dcm_dict)
                elif self.scanner_ID == "45504":
                    self.sorted_dcm_dict = self.Sort_Dicom_Dict_Elition_X_PHT(self.dcm_dict)
               
                #archive DICOMs in a sorted format
                if self.archive == True:
                    self.Archive_Images(self.sorted_dcm_dict, self.archive_directory)
                #produce PNGs of magnitude images
                if self.figures == True:
                    self.Archive_PNGs(self.sorted_dcm_dict)
                #Get the default thresholds used for producing the mask for the coil selected
                #If different thresholds are required this may indicate a fault.
                self.lower_threshold = self.coil_dict[self.scanner_ID][self.coil_name]["lower_threshold"]
                self.upper_threshold = self.coil_dict[self.scanner_ID][self.coil_name]["upper_threshold"]
                
        def Select_Directory(self, initial_dir, window_title):
            """
            Parameters
            ----------
            initial_dir: str
                Path at which the file dialogue opens
            window_title: str
                Title of the filedialogue window.
                
            Returns
            -------
            user_selected_directory : str
            
            Uses tkinter filedialog.  Asks the user to select a directory and
            returns the user_selected_directory.
            """
            directory_window = tk.Toplevel()
            directory_window.wm_attributes('-topmost', 1)
            directory_window.withdraw()
            user_selected_directory =  filedialog.askdirectory(parent=directory_window, 
                                                               initialdir=initial_dir, 
                                                               title=window_title)
            directory_window.destroy()
            return user_selected_directory
    
    
        def Get_Files(self, starting_dir):
            """
            Parameters
            ----------
            starting_dir: str
                Every file not called VERSION inside 
                starting_dir will be assumed to be a dicom of interest.
            
            Returns
            -------
            paths : list of str
                list of dicom_paths inside starting_dir.
            """
            paths = []
             
            for dirName, subdirList, fileList in os.walk(starting_dir):
                #individual channel data
                for filename in fileList:
                    if filename != "VERSION":
                        paths.append(os.path.join(dirName,filename))
                    
            return paths
        
        def Get_Dicom_Dict(self, dicom_paths):
            """
            Parameters
            ----------
            dicom_paths : list of strings
                List of all dicom paths
            Returns
            -------
            dcm_dict : dict
                Dictionary of dicoms of the form {<path>:<dicom>}
            """
            dcm_dict = {}
            for dicom_path in dicom_paths:
                dcm = pydicom.read_file(dicom_path)
                if dcm.SeriesDescription in self.SNR_sequence_names:
                    if dcm.DeviceSerialNumber in ["183188","202541","183109"]:
                        #ensure only Non distortion corrected images are used
                        if dcm.SeriesDescription[-2:] == "ND":
                            dcm_dict[dicom_path] = dcm
                    else:
                        dcm_dict[dicom_path] = dcm
            return dcm_dict
        
        def Get_Coil_Name(self, dcm_dict):
            """
            Parameters
            ----------
            dcm_dict : dict
                Dictionary of dicoms of the form {<path>:<dicom>}
            
            
            Initialises the scanner ID, the coil name, the number of elements 
            the combined elemet image is produced from and the scanner name. 
            It is assumed that all the dicoms have been acuired on the same 
            scanner and coil.  
            
            """
            path = list(dcm_dict.keys())[0]
            self.scanner_ID = dcm_dict[path].DeviceSerialNumber
            self.categories = list(self.coil_dict[self.scanner_ID].keys())
            top = tk.Toplevel()
            top.title("Select Coil Name")
            top.geometry('200x100')
            top.wm_attributes('-topmost', 1)
            dcm_category = tk.StringVar()
            dcm_category.set(self.categories[0])
            drop = tk.OptionMenu(top, dcm_category, *self.categories).pack()
            proceed = tk.IntVar()
            btn = tk.Button(top, text="OK", command=lambda: proceed.set(1)).pack()
            top.wait_variable(proceed)
            self.coil_name = dcm_category.get()
            self.n_elements = self.coil_dict[self.scanner_ID][self.coil_name]["n_elements"]
            self.scanner_name = self.scanner_ID_dict[self.scanner_ID]
            top.destroy()
            
        def Sort_Dicom_Dict_Ambition_X_PHT(self, dcm_dict):
            sorted_dcm_dict={}
            n_elements = self.coil_dict[self.scanner_ID][self.coil_name]["n_elements"]
            n_slices = self.coil_dict[self.scanner_ID][self.coil_name]["n_slices"]
            for path in dcm_dict:
                dcm = dcm_dict[path]
                
                
                if dcm.AcquisitionDate not in sorted_dcm_dict.keys():
                    sorted_dcm_dict[dcm.AcquisitionDate] = {}
                if dcm.SeriesTime not in sorted_dcm_dict[dcm.AcquisitionDate].keys():
                    sorted_dcm_dict[dcm.AcquisitionDate][dcm.SeriesTime] = {}   
                if dcm.AcquisitionTime not in sorted_dcm_dict[dcm.AcquisitionDate][dcm.SeriesTime].keys():
                    sorted_dcm_dict[dcm.AcquisitionDate][dcm.SeriesTime][dcm.AcquisitionTime] = {}
                    sorted_dcm_dict[dcm.AcquisitionDate][dcm.SeriesTime][dcm.AcquisitionTime]["DelRec"] = {}
                    sorted_dcm_dict[dcm.AcquisitionDate][dcm.SeriesTime][dcm.AcquisitionTime]["Combined"] = {}
                    sorted_dcm_dict[dcm.AcquisitionDate][dcm.SeriesTime][dcm.AcquisitionTime]["Combined"][1] = {}
                    sorted_dcm_dict[dcm.AcquisitionDate][dcm.SeriesTime][dcm.AcquisitionTime]["DelRec"]["unsorted"] = {}
                if dcm.ProtocolName[0:6] =="DelRec":
                    #individual elements
                    sorted_dcm_dict[dcm.AcquisitionDate][dcm.SeriesTime][dcm.AcquisitionTime]["DelRec"]["unsorted"][dcm.InstanceNumber] = {"path":path, "dcm":dcm}
                    
            for acq_date in sorted_dcm_dict:
                for series_time in sorted_dcm_dict[acq_date]:
                    for acq_time in sorted_dcm_dict[acq_date][series_time]:
                        #assume 1 slice
                        n_slices = 1
                        n_elements = int(len(sorted_dcm_dict[acq_date][series_time][acq_time]["DelRec"]["unsorted"])/n_slices)
                        if self.coil_dict[self.scanner_ID][self.coil_name]["n_elements"] != n_elements:
                            raise Exception("Incorrect number of elements for " + self.coil_name + str(n_elements) +" calculated")
                        if self.coil_dict[self.scanner_ID][self.coil_name]["n_slices"] != n_slices:
                            raise Exception("Incorrect number of slices for " + self.coil_name + str(n_slices) +" calculated")
                        for  instance_n in sorted_dcm_dict[acq_date][series_time][acq_time]["DelRec"]["unsorted"]:
                            slice_n = 1 
                            element_n = int(instance_n % n_elements)+1
                            if slice_n not in sorted_dcm_dict[acq_date][series_time][acq_time]["DelRec"].keys():
                                sorted_dcm_dict[acq_date][series_time][acq_time]["DelRec"][slice_n] = {}
                            dcm = sorted_dcm_dict[acq_date][series_time][acq_time]["DelRec"]["unsorted"][instance_n]["dcm"]
                            
                            try:
                                SoS_array = np.add(SoS_array,np.square(dcm.pixel_array.astype(float)))
                            except UnboundLocalError:
                                SoS_array = np.square(dcm.pixel_array.astype(float))
                            
                            sorted_dcm_dict[acq_date][series_time][acq_time]["DelRec"][slice_n][element_n] = sorted_dcm_dict[acq_date][series_time][acq_time]["DelRec"]["unsorted"][instance_n]            
                        del sorted_dcm_dict[acq_date][series_time][acq_time]["DelRec"]["unsorted"]
                        sorted_dcm_dict[acq_date][series_time][acq_time]["Combined"][1]["pixel_array"] = np.sqrt(SoS_array)
                        del SoS_array
            
            
            
            return sorted_dcm_dict
        
        
        def Sort_Dicom_Dict_Elition_X_PHT(self, dcm_dict):
            sorted_dcm_dict={}
            n_elements = self.coil_dict[self.scanner_ID][self.coil_name]["n_elements"]
            n_slices = self.coil_dict[self.scanner_ID][self.coil_name]["n_slices"]
            for path in dcm_dict:
                dcm = dcm_dict[path]
                #t1 = Tag(0x7a1103e)
                t1 = Tag(0x0200013)
                acquisition_number = dcm[t1].value
                if dcm.SeriesDate not in sorted_dcm_dict.keys():
                    sorted_dcm_dict[dcm.SeriesDate] = {}
                if dcm.SeriesTime not in sorted_dcm_dict[dcm.SeriesDate].keys():
                    sorted_dcm_dict[dcm.SeriesDate][dcm.SeriesTime] = {}
                    sorted_dcm_dict[dcm.SeriesDate][dcm.SeriesTime]["unsorted"] = {}
                   
                if dcm.ProtocolName[0:6] =="DelRec":
                    #individual elements
                    sorted_dcm_dict[dcm.SeriesDate][dcm.SeriesTime]["unsorted"][acquisition_number] = {"path":path, "dcm":dcm}
                    
            for acq_date in sorted_dcm_dict:
                for series_time in sorted_dcm_dict[acq_date]:
                     #assume 1 slice
                    n_slices = 1
                    n_elements = int(len(sorted_dcm_dict[acq_date][series_time]["unsorted"])/(2*n_slices))
                    if self.coil_dict[self.scanner_ID][self.coil_name]["n_elements"] != n_elements:
                        raise Exception("Incorrect number of elements for " + self.coil_name + ", " + str(n_elements) +" calculated")
                    if self.coil_dict[self.scanner_ID][self.coil_name]["n_slices"] != n_slices:
                        raise Exception("Incorrect number of slices for " + self.coil_name + str(n_slices) +" calculated")
                    if len(sorted_dcm_dict[acq_date][series_time]["unsorted"]) == 2:
                        #single element
                        for image_n in sorted_dcm_dict[acq_date][series_time]["unsorted"]:
                            sorted_dcm_dict[acq_date][series_time][image_n] = {}
                            sorted_dcm_dict[acq_date][series_time][image_n]["DelRec"] = {}
                            sorted_dcm_dict[acq_date][series_time][image_n]["Combined"] = {}
                            sorted_dcm_dict[acq_date][series_time][image_n]["Combined"][1] = sorted_dcm_dict[dcm.SeriesDate][dcm.SeriesTime]["unsorted"][image_n]
                            sorted_dcm_dict[acq_date][series_time][image_n]["DelRec"][1] = {}
                            sorted_dcm_dict[acq_date][series_time][image_n]["DelRec"][1][1] = sorted_dcm_dict[dcm.SeriesDate][dcm.SeriesTime]["unsorted"][image_n]
                        
                        del sorted_dcm_dict[acq_date][series_time]["unsorted"]
                    else:
                        SoS_arrays = {}
                        for image_n in sorted_dcm_dict[acq_date][series_time]["unsorted"]:
                            slice_n = 1 
                            element_n = int(image_n % n_elements)+1
                            repeat_n = int((image_n-1)/n_elements)
                            
                            if repeat_n not in sorted_dcm_dict[acq_date][series_time].keys():
                                sorted_dcm_dict[acq_date][series_time][repeat_n] = {}
                                sorted_dcm_dict[acq_date][series_time][repeat_n]["DelRec"] = {}
                                sorted_dcm_dict[acq_date][series_time][repeat_n]["Combined"] = {}
                                sorted_dcm_dict[acq_date][series_time][repeat_n]["DelRec"][1] = {}
                                sorted_dcm_dict[acq_date][series_time][repeat_n]["Combined"][1] = {}
                            
                            sorted_dcm_dict[acq_date][series_time][repeat_n]["DelRec"][1][element_n] = sorted_dcm_dict[acq_date][series_time]["unsorted"][image_n]   
                            
                                                            
                            dcm = sorted_dcm_dict[acq_date][series_time]["unsorted"][image_n]["dcm"]
                            
                            try:
                                SoS_arrays[repeat_n] = np.add(SoS_arrays[repeat_n],np.square(dcm.pixel_array.astype(float)))
                            except KeyError:
                                SoS_arrays[repeat_n] = np.square(dcm.pixel_array.astype(float))
                            
                            sorted_dcm_dict[acq_date][series_time][repeat_n]["DelRec"][slice_n][element_n] = sorted_dcm_dict[acq_date][series_time]["unsorted"][image_n]            
                        
                        for repeat_n in SoS_arrays:
                            sorted_dcm_dict[acq_date][series_time][repeat_n]["Combined"][1]["pixel_array"] = np.sqrt(SoS_arrays[repeat_n])
                        del sorted_dcm_dict[acq_date][series_time]["unsorted"]
                        del SoS_arrays
            
            
            
            return sorted_dcm_dict
        
        
        
        
        def Sort_Dicom_Dict_Siemens(self, dcm_dict):
            #LEGACY SORTING METHOD FOR SIEMENS SCANNERS
            sorted_dcm_dict={}
            
            for path in dcm_dict:
                dcm = dcm_dict[path]
                
                #a=dcm.dir()
                if dcm.AcquisitionDate not in sorted_dcm_dict.keys():
                    sorted_dcm_dict[dcm.AcquisitionDate] = {}  
                """Series times of the Siemens individual elements and combined images are <10ms different. Need to check not already a series with a series time <1 second different """
                
                if dcm.SeriesTime not in sorted_dcm_dict[dcm.AcquisitionDate].keys():
                    test_time = str(int(dcm.SeriesTime) + 1).zfill(6)
                    if test_time not in sorted_dcm_dict[dcm.AcquisitionDate].keys():
                        test_time = str(int(dcm.SeriesTime) - 1).zfill(6)
                        if test_time not in sorted_dcm_dict[dcm.AcquisitionDate].keys():
                            series_time = dcm.SeriesTime
                            sorted_dcm_dict[dcm.AcquisitionDate][series_time] = {}  
                            sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"] = {}
                            sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"]["DelRec"] = {}
                            sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"]["Combined"] = {}
                            sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"]["DelRec"]["unsorted"] = {}
                        else:
                            series_time = test_time
                    else:
                        series_time = test_time
                else:
                    series_time = dcm.SeriesTime
                           
                
                #Image Number (InstanceNumber) is slice number
                if "NORM" in dcm.ImageType._list:
                    sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"]["Combined"][dcm.InstanceNumber] = {"path":path, "dcm":dcm}
                else:
                    #individual elements
                    if dcm.InstanceNumber not in sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"]["DelRec"]["unsorted"]:
                        sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"]["DelRec"]["unsorted"][dcm.InstanceNumber] = {}
                    sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"]["DelRec"]["unsorted"][dcm.InstanceNumber][dcm.ContentTime] = {"path":path, "dcm":dcm}
     
                    
            for acq_date in sorted_dcm_dict:
                for series_time in sorted_dcm_dict[acq_date]:
                    n_slices = len(sorted_dcm_dict[acq_date][series_time]["unsorted"]["Combined"])
                    n_elements = len(sorted_dcm_dict[acq_date][series_time]["unsorted"]["DelRec"]["unsorted"][1])
                    if self.coil_dict[self.scanner_ID][self.coil_name]["n_elements"] != n_elements:
                        raise Exception("Incorrect number of elements for " + self.coil_name + str(n_elements) +" calculated")
                    if self.coil_dict[self.scanner_ID][self.coil_name]["n_slices"] != n_slices:
                        raise Exception("Incorrect number of slices for " + self.coil_name + str(n_slices) +" calculated")
                    for  slice_n in sorted_dcm_dict[acq_date][series_time]["unsorted"]["DelRec"]["unsorted"]:
                        if slice_n not in sorted_dcm_dict[acq_date][series_time]["unsorted"]["DelRec"].keys():
                            sorted_dcm_dict[acq_date][series_time]["unsorted"]["DelRec"][slice_n] = {}
                        
                        acq_times = [*sorted_dcm_dict[acq_date][series_time]["unsorted"]["DelRec"]["unsorted"][slice_n]]
                        acq_times.sort()
                        for i in range(len(acq_times)):
                            acq_time = acq_times[i]
                            element_n = 1 + i
                            sorted_dcm_dict[acq_date][series_time]["unsorted"]["DelRec"][slice_n][element_n] = sorted_dcm_dict[acq_date][series_time]["unsorted"]["DelRec"]["unsorted"][slice_n][acq_time]          
                    del sorted_dcm_dict[acq_date][series_time]["unsorted"]["DelRec"]["unsorted"]
            
            for acq_date in sorted_dcm_dict:
                series_times = list(sorted_dcm_dict[acq_date].keys())
                series_times.sort()
                if len(series_times) % 2 != 0:
                    del series_times[-1]
                for i in range(len(series_times)):
                    current_series_time = series_times[i]
                    if i % 2 == 0:
                        new_series_time = current_series_time
                        acq_number = "0"
                    else:
                        new_series_time = series_times[i-1]
                        acq_number = "1"
                    
                    sorted_dcm_dict[acq_date][new_series_time][current_series_time] = sorted_dcm_dict[acq_date][current_series_time]["unsorted"]
                    
                    if acq_number == "0":
                        del sorted_dcm_dict[acq_date][current_series_time]["unsorted"]
                    else:
                        del sorted_dcm_dict[acq_date][current_series_time]
            
            
            
            return sorted_dcm_dict
        
        def Sort_Dicom_Dict_Siemens_Avanto_RBH_PHT(self, dcm_dict):
            """
            Parameters
            ----------
            dcm_dict : dict
                Dictionary of dicoms of the form {<path>:<dicom>}
            
            Returns
            -------
            sorted_dcm_dict : dict
                Dictionary of dicoms of the form {<Date>:{<Series Time>:
                {<repeat_number>:{"DelRec":{<Slice number>:{<element number>:
                                        {"path":<path>, "dcm":<DICOM>}}},
                                "Combined":{<Slice number>:
                                        {"path":<path>, "dcm":<DICOM>}}}}}}
            
            Sorts DICOMs exported from RBH Sielems Sola (MR3).  Images are sorted by:
                -Acquisition Date: all images should be acquired on the same date
                -Series Time: There should be 2 repeats for each series time.  Each repeat should have corresponding combined and individual element images
                -Repeat number: should be 2 repeats
                -Image type: Sorts images into combined("Combined" and individual element images("DelRec")
                -Slice number:  Legacy from old analysis when multiple slices were acquired per coil.  Now only a single slice is used
                -Element number:  Only a key for DelRec
            """
            #dictionary to be populated
            sorted_dcm_dict={}
            for path in dcm_dict:
                dcm = dcm_dict[path]
                if dcm.AcquisitionDate not in sorted_dcm_dict.keys():
                    sorted_dcm_dict[dcm.AcquisitionDate] = {}  
                """Series times of the Siemens individual elements and combined images are <10ms different. Need to check not already a series with a series time <1 second different """
                              
                if dcm.SeriesTime not in sorted_dcm_dict[dcm.AcquisitionDate].keys():
                    test_time = str(int(dcm.SeriesTime) + 1).zfill(6)
                    if test_time not in sorted_dcm_dict[dcm.AcquisitionDate].keys():
                        test_time = str(int(dcm.SeriesTime) - 1).zfill(6)
                        if test_time not in sorted_dcm_dict[dcm.AcquisitionDate].keys():
                            series_time = dcm.SeriesTime
                            sorted_dcm_dict[dcm.AcquisitionDate][series_time] = {}  
                            sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"] = {}
                        else:
                            series_time = test_time
                    else:
                        series_time = test_time
                else:
                    series_time = dcm.SeriesTime
                
                
                
                if series_time not in sorted_dcm_dict[dcm.AcquisitionDate].keys():
                    sorted_dcm_dict[dcm.AcquisitionDate][series_time] = {}
                    sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"] = {}
                    
                if dcm.SeriesNumber not in sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"]:
                    #seperate uncombined and combined by series number
                    sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"][dcm.SeriesNumber] = {}
                if dcm.InstanceNumber not in sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"][dcm.SeriesNumber]:
                    sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"][dcm.SeriesNumber][dcm.InstanceNumber] = {}
                
                #Initially sort DICOMS by Series Number and Instance number 
                #these are used to assess if the acquisition is a combined acquisition/ delrec
                #and what the element number is
                
                
                if dcm.InstanceCreationTime not in sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"][dcm.SeriesNumber][dcm.InstanceNumber]:
                    sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"][dcm.SeriesNumber][dcm.InstanceNumber][dcm.InstanceCreationTime] = {}
                
                sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"][dcm.SeriesNumber][dcm.InstanceNumber][dcm.InstanceCreationTime] = {"path":path, "dcm":dcm}
                
     
            n_elements = self.coil_dict[self.scanner_ID][self.coil_name]["n_elements"]
            n_slices = self.coil_dict[self.scanner_ID][self.coil_name]["n_slices"]
            for acq_date in sorted_dcm_dict:
                for series_time in sorted_dcm_dict[acq_date]:
                    for series_number in sorted_dcm_dict[acq_date][series_time]["unsorted"]:
                        repeat_ns=list(sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number].keys())
                        if len(sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number][repeat_ns[0]]) == n_slices:
                            #combined_elements_image
                            image_type = "Combined"
                            for repeat_number in sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number]:
                                repeat_number_int = int(repeat_number)
                                if repeat_number_int not in sorted_dcm_dict[acq_date][series_time]:
                                    sorted_dcm_dict[acq_date][series_time][repeat_number_int] = {}
                                if image_type not in sorted_dcm_dict[acq_date][series_time][repeat_number_int]:
                                    sorted_dcm_dict[acq_date][series_time][repeat_number_int][image_type] = {}
                                    #1 slice
                                    sorted_dcm_dict[acq_date][series_time][repeat_number_int][image_type][1] = {}
                                instance_times = list(sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number][repeat_number].keys())
                                if len(instance_times) != 1:
                                    print("Sort Failed")
                                sorted_dcm_dict[acq_date][series_time][repeat_number_int][image_type][1] = sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number][repeat_number][instance_times[0]]
                        elif len(sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number][repeat_ns[0]]) == n_slices*n_elements:
                            #number of del rec images is elements*slices*2   
                            #individual_elements_image
                            image_type = "DelRec"
                            
                            for repeat_number in sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number]:
                                repeat_number_int = int(repeat_number)
                                if repeat_number_int not in sorted_dcm_dict[acq_date][series_time]:
                                    sorted_dcm_dict[acq_date][series_time][repeat_number_int] = {}
                                if image_type not in sorted_dcm_dict[acq_date][series_time][repeat_number_int]:
                                    sorted_dcm_dict[acq_date][series_time][repeat_number_int][image_type] = {}
                                    #1 slice
                                    sorted_dcm_dict[acq_date][series_time][repeat_number_int][image_type][1] = {}
                                
                                instance_times = sorted(list(sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number][repeat_number].keys()),key=float)
                                for instance_time in instance_times:
                                    element_number = instance_times.index(instance_time)+1
                                    if element_number not in sorted_dcm_dict[acq_date][series_time][repeat_number][image_type][1]:
                                        sorted_dcm_dict[acq_date][series_time][repeat_number][image_type][1][element_number] = {}
                                    sorted_dcm_dict[acq_date][series_time][repeat_number][image_type][1][element_number] = sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number][repeat_number][instance_time]
                   
                    del sorted_dcm_dict[acq_date][series_time]["unsorted"]            
            return sorted_dcm_dict
        
        def Sort_Dicom_Dict_Siemens_Avanto_DCH(self, dcm_dict):
            """
            Parameters
            ----------
            dcm_dict : dict
                Dictionary of dicoms of the form {<path>:<dicom>}
            
            Returns
            -------
            sorted_dcm_dict : dict
                Dictionary of dicoms of the form {<Date>:{<Series Time>:
                {<repeat_number>:{"DelRec":{<Slice number>:{<element number>:
                                        {"path":<path>, "dcm":<DICOM>}}},
                                "Combined":{<Slice number>:
                                        {"path":<path>, "dcm":<DICOM>}}}}}}
            
            Sorts DICOMs exported from RBH Sielems Sola (MR3).  Images are sorted by:
                -Acquisition Date: all images should be acquired on the same date
                -Series Time: There should be 2 repeats for each series time.  Each repeat should have corresponding combined and individual element images
                -Repeat number: should be 2 repeats
                -Image type: Sorts images into combined("Combined" and individual element images("DelRec")
                -Slice number:  Legacy from old analysis when multiple slices were acquired per coil.  Now only a single slice is used
                -Element number:  Only a key for DelRec
            """
            #dictionary to be populated
            sorted_dcm_dict={}
            for path in dcm_dict:
                dcm = dcm_dict[path]
                if dcm.AcquisitionDate not in sorted_dcm_dict.keys():
                    sorted_dcm_dict[dcm.AcquisitionDate] = {}  
                """Series times of the Siemens individual elements and combined images are <10ms different. Need to check not already a series with a series time <1 second different """
                              
                if str(int(float(dcm.SeriesTime))).zfill(6) not in sorted_dcm_dict[dcm.AcquisitionDate].keys():
                    test_time = str(int(float(dcm.SeriesTime)) + 1).zfill(6)
                    if test_time not in sorted_dcm_dict[dcm.AcquisitionDate].keys():
                        test_time = str(int(float(dcm.SeriesTime)) - 1).zfill(6)
                        if test_time not in sorted_dcm_dict[dcm.AcquisitionDate].keys():
                            series_time = str(int(float(dcm.SeriesTime))).zfill(6)
                        else:
                            series_time = test_time
                    else:
                        series_time = test_time
                else:
                    series_time = str(int(float(dcm.SeriesTime))).zfill(6)
                
                
                
                if series_time not in sorted_dcm_dict[dcm.AcquisitionDate].keys():
                    sorted_dcm_dict[dcm.AcquisitionDate][series_time] = {}
                    sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"] = {}
                    
                if dcm.SeriesNumber not in sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"]:
                    #seperate uncombined and combined by series number
                    sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"][dcm.SeriesNumber] = {}
                if dcm.InstanceNumber not in sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"][dcm.SeriesNumber]:
                    sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"][dcm.SeriesNumber][dcm.InstanceNumber] = {}
                
                #Initially sort DICOMS by Series Number and Instance number 
                #these are used to assess if the acquisition is a combined acquisition/ delrec
                #and what the element number is
                
                
                if dcm.InstanceCreationTime not in sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"][dcm.SeriesNumber][dcm.InstanceNumber]:
                    sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"][dcm.SeriesNumber][dcm.InstanceNumber][dcm.InstanceCreationTime] = {}
                
                sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"][dcm.SeriesNumber][dcm.InstanceNumber][dcm.InstanceCreationTime] = {"path":path, "dcm":dcm}
                
     
            n_elements = self.coil_dict[self.scanner_ID][self.coil_name]["n_elements"]
            n_slices = self.coil_dict[self.scanner_ID][self.coil_name]["n_slices"]
            for acq_date in sorted_dcm_dict:
                for series_time in sorted_dcm_dict[acq_date]:
                    for series_number in sorted_dcm_dict[acq_date][series_time]["unsorted"]:
                        repeat_ns=list(sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number].keys())
                        if len(sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number][repeat_ns[0]]) == n_slices:
                            #combined_elements_image
                            image_type = "Combined"
                            for repeat_number in sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number]:
                                repeat_number_int = int(repeat_number)
                                if repeat_number_int not in sorted_dcm_dict[acq_date][series_time]:
                                    sorted_dcm_dict[acq_date][series_time][repeat_number_int] = {}
                                if image_type not in sorted_dcm_dict[acq_date][series_time][repeat_number_int]:
                                    sorted_dcm_dict[acq_date][series_time][repeat_number_int][image_type] = {}
                                    #1 slice
                                    sorted_dcm_dict[acq_date][series_time][repeat_number_int][image_type][1] = {}
                                instance_times = list(sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number][repeat_number].keys())
                                if len(instance_times) != 1:
                                    print("Sort Failed")
                                sorted_dcm_dict[acq_date][series_time][repeat_number_int][image_type][1] = sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number][repeat_number][instance_times[0]]
                        elif len(sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number][repeat_ns[0]]) == n_slices*n_elements:
                            #number of del rec images is elements*slices*2   
                            #individual_elements_image
                            image_type = "DelRec"
                            
                            for repeat_number in sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number]:
                                repeat_number_int = int(repeat_number)
                                if repeat_number_int not in sorted_dcm_dict[acq_date][series_time]:
                                    sorted_dcm_dict[acq_date][series_time][repeat_number_int] = {}
                                if image_type not in sorted_dcm_dict[acq_date][series_time][repeat_number_int]:
                                    sorted_dcm_dict[acq_date][series_time][repeat_number_int][image_type] = {}
                                    #1 slice
                                    sorted_dcm_dict[acq_date][series_time][repeat_number_int][image_type][1] = {}
                                
                                instance_times = sorted(list(sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number][repeat_number].keys()),key=float)
                                for instance_time in instance_times:
                                    element_number = instance_times.index(instance_time)+1
                                    if element_number not in sorted_dcm_dict[acq_date][series_time][repeat_number][image_type][1]:
                                        sorted_dcm_dict[acq_date][series_time][repeat_number][image_type][1][element_number] = {}
                                    sorted_dcm_dict[acq_date][series_time][repeat_number][image_type][1][element_number] = sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number][repeat_number][instance_time]
                   
                    del sorted_dcm_dict[acq_date][series_time]["unsorted"]            
            return sorted_dcm_dict

        
        def Sort_Dicom_Dict_Siemens_Sola_RBH(self, dcm_dict):
            """
            Parameters
            ----------
            dcm_dict : dict
                Dictionary of dicoms of the form {<path>:<dicom>}
            
            Returns
            -------
            sorted_dcm_dict : dict
                Dictionary of dicoms of the form {<Date>:{<Series Time>:
                {<repeat_number>:{"DelRec":{<Slice number>:{<element number>:
                                        {"path":<path>, "dcm":<DICOM>}}},
                                "Combined":{<Slice number>:
                                        {"path":<path>, "dcm":<DICOM>}}}}}}
            
            Sorts DICOMs exported from RBH Sielems Sola (MR3).  Images are sorted by:
                -Acquisition Date: all images should be acquired on the same date
                -Series Time: There should be 2 repeats for each series time.  Each repeat should have corresponding combined and individual element images
                -Repeat number: should be 2 repeats
                -Image type: Sorts images into combined("Combined" and individual element images("DelRec")
                -Slice number:  Legacy from old analysis when multiple slices were acquired per coil.  Now only a single slice is used
                -Element number:  Only a key for DelRec
            """
            #dictionary to be populated
            sorted_dcm_dict={}
            for path in dcm_dict:
                dcm = dcm_dict[path]
                
                #a=dcm.dir()
                if dcm.AcquisitionDate not in sorted_dcm_dict.keys():
                    sorted_dcm_dict[dcm.AcquisitionDate] = {}  
                """Series times of the Siemens individual elements and combined images are <10ms different. Need to check not already a series with a series time <1 second different """
                              
                if dcm.SeriesTime not in sorted_dcm_dict[dcm.AcquisitionDate].keys():
                    test_time = str(int(dcm.SeriesTime) + 1).zfill(6)
                    if test_time not in sorted_dcm_dict[dcm.AcquisitionDate].keys():
                        test_time = str(int(dcm.SeriesTime) - 1).zfill(6)
                        if test_time not in sorted_dcm_dict[dcm.AcquisitionDate].keys():
                            series_time = dcm.SeriesTime
                            sorted_dcm_dict[dcm.AcquisitionDate][series_time] = {}  
                            sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"] = {}
                            sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"]["DelRec"] = {}
                            sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"]["Combined"] = {}
                            sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"]["DelRec"]["unsorted"] = {}
                        else:
                            series_time = test_time
                    else:
                        series_time = test_time
                else:
                    series_time = dcm.SeriesTime
                
                
                
                if series_time not in sorted_dcm_dict[dcm.AcquisitionDate].keys():
                    sorted_dcm_dict[dcm.AcquisitionDate][series_time] = {}
                    sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"] = {}
                    
                if dcm.SeriesNumber not in sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"]:
                    #seperate uncombined and combined by series number
                    sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"][dcm.SeriesNumber] = {}
                if dcm.InstanceNumber not in sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"][dcm.SeriesNumber]:
                    sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"][dcm.SeriesNumber][dcm.InstanceNumber] = {}
                
                #Initially sort DICOMS by Series Number and Instance number 
                #these are used to assess if the acquisition is a combined acquisition/ delrec
                #and what the element number is
                sorted_dcm_dict[dcm.AcquisitionDate][series_time]["unsorted"][dcm.SeriesNumber][dcm.InstanceNumber] = {"path":path, "dcm":dcm}
                
     
            n_elements = self.coil_dict[self.scanner_ID][self.coil_name]["n_elements"]
            n_slices = self.coil_dict[self.scanner_ID][self.coil_name]["n_slices"]
            for acq_date in sorted_dcm_dict:
                for series_time in sorted_dcm_dict[acq_date]:
                    for series_number in sorted_dcm_dict[acq_date][series_time]["unsorted"]:
                        if len(sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number]) == 2*n_slices:
                            #combined_elements_image
                            image_type = "Combined"
                            for image_number in sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number]:
                                repeat_number = int(image_number)
                                if repeat_number not in sorted_dcm_dict[acq_date][series_time]:
                                    sorted_dcm_dict[acq_date][series_time][repeat_number] = {}
                                if image_type not in sorted_dcm_dict[acq_date][series_time][repeat_number]:
                                    sorted_dcm_dict[acq_date][series_time][repeat_number][image_type] = {}
                                    #1 slice
                                    sorted_dcm_dict[acq_date][series_time][repeat_number][image_type][1] = {}
                                sorted_dcm_dict[acq_date][series_time][repeat_number][image_type][1] = sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number][image_number]
                        elif len(sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number]) == 2*n_slices*n_elements:
                            #number of del rec images is elements*slices*2   
                            #individual_elements_image
                            image_type = "DelRec"
                            
                            for image_number in sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number]:
                                
                                element_number = image_number % n_elements
                                if element_number == 0:
                                    element_number = n_elements
                                if element_number == image_number:
                                    repeat_number = 1
                                else:
                                    repeat_number = 2
                                if repeat_number not in sorted_dcm_dict[acq_date][series_time]:
                                    sorted_dcm_dict[acq_date][series_time][repeat_number] = {}
                                if image_type not in sorted_dcm_dict[acq_date][series_time][repeat_number]:
                                    sorted_dcm_dict[acq_date][series_time][repeat_number][image_type] = {}
                                    #1 slice
                                    sorted_dcm_dict[acq_date][series_time][repeat_number][image_type][1] = {}
                                if element_number not in sorted_dcm_dict[acq_date][series_time][repeat_number][image_type][1]:
                                    sorted_dcm_dict[acq_date][series_time][repeat_number][image_type][1][element_number] = {}
                                sorted_dcm_dict[acq_date][series_time][repeat_number][image_type][1][element_number] = sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number][image_number]
                   
                    del sorted_dcm_dict[acq_date][series_time]["unsorted"]            
            return sorted_dcm_dict

        def Sort_Dicom_Dict_Siemens_Sola_DCH(self, dcm_dict):
            """
            Parameters
            ----------
            dcm_dict : dict
                Dictionary of dicoms of the form {<path>:<dicom>}
            
            Returns
            -------
            sorted_dcm_dict : dict
                Dictionary of dicoms of the form {<Date>:{<Series Time>:
                {<repeat_number>:{"DelRec":{<Slice number>:{<element number>:
                                        {"path":<path>, "dcm":<DICOM>}}},
                                "Combined":{<Slice number>:
                                        {"path":<path>, "dcm":<DICOM>}}}}}}
            
            Sorts DICOMs exported from DCH Sielems Sola (MR2).  Images are sorted by:
                -Acquisition Date: all images should be acquired on the same date
                -Series Time: There should be 2 repeats for each series time.  Each repeat should have corresponding combined and individual element images
                -Repeat number: should be 2 repeats
                -Image type: Sorts images into combined("Combined" and individual element images("DelRec")
                -Slice number:  Legacy from old analysis when multiple slices were acquired per coil.  Now only a single slice is used
                -Element number:  Only a key for DelRec
            """
            sorted_dcm_dict={}
            
            for path in dcm_dict:
                dcm = dcm_dict[path]
                
                
                if dcm.SeriesDate not in sorted_dcm_dict.keys():
                    sorted_dcm_dict[dcm.SeriesDate] = {}  
                """Series times of the Siemens individual elements and combined images are <10ms different. Need to check not already a series with a series time <1 second different """
                
                #a=dcm.dir("")
                #print(dcm)
                
                if str(int(float(dcm.SeriesTime))) not in sorted_dcm_dict[dcm.SeriesDate].keys():
                    test_time = str(int(float(dcm.SeriesTime)) + 1)
                    if test_time not in sorted_dcm_dict[dcm.SeriesDate].keys():
                        test_time = str(int(float(dcm.SeriesTime)) - 1)
                        if test_time not in sorted_dcm_dict[dcm.SeriesDate].keys():
                            series_time = str(int(float(dcm.SeriesTime)))
                            sorted_dcm_dict[dcm.SeriesDate][series_time] = {}  
                            sorted_dcm_dict[dcm.SeriesDate][series_time]["unsorted"] = {}
                            sorted_dcm_dict[dcm.SeriesDate][series_time]["unsorted"]["DelRec"] = {}
                            sorted_dcm_dict[dcm.SeriesDate][series_time]["unsorted"]["Combined"] = {}
                            sorted_dcm_dict[dcm.SeriesDate][series_time]["unsorted"]["DelRec"]["unsorted"] = {}
                        else:
                            series_time = test_time
                    else:
                        series_time = test_time
                else:
                    series_time = str(int(float(dcm.SeriesTime)))
                
                
                
                if series_time not in sorted_dcm_dict[dcm.SeriesDate].keys():
                    sorted_dcm_dict[dcm.SeriesDate][series_time] = {}
                    sorted_dcm_dict[dcm.SeriesDate][series_time]["unsorted"] = {}
                    
                if dcm.SeriesNumber not in sorted_dcm_dict[dcm.SeriesDate][series_time]["unsorted"]:
                    #seperate uncombined and combined by series number
                    sorted_dcm_dict[dcm.SeriesDate][series_time]["unsorted"][dcm.SeriesNumber] = {}
                if dcm.InstanceNumber not in sorted_dcm_dict[dcm.SeriesDate][series_time]["unsorted"][dcm.SeriesNumber]:
                    sorted_dcm_dict[dcm.SeriesDate][series_time]["unsorted"][dcm.SeriesNumber][dcm.InstanceNumber] = {}
                #Initially sort DICOMS by Series Number and Instance number 
                #these are used to assess if the acquisition is a combined acquisition/ delrec
                #and what the element number is
                sorted_dcm_dict[dcm.SeriesDate][series_time]["unsorted"][dcm.SeriesNumber][dcm.InstanceNumber] = {"path":path, "dcm":dcm}
                
     
            n_elements = self.coil_dict[self.scanner_ID][self.coil_name]["n_elements"]
            n_slices = self.coil_dict[self.scanner_ID][self.coil_name]["n_slices"]
            for acq_date in sorted_dcm_dict:
                for series_time in sorted_dcm_dict[acq_date]:
                    for series_number in sorted_dcm_dict[acq_date][series_time]["unsorted"]:
                        if len(sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number]) == 2*n_slices:
                            #combined_elements_image
                            image_type = "Combined"
                            for image_number in sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number]:
                                repeat_number = int(image_number)
                                if repeat_number not in sorted_dcm_dict[acq_date][series_time]:
                                    sorted_dcm_dict[acq_date][series_time][repeat_number] = {}
                                if image_type not in sorted_dcm_dict[acq_date][series_time][repeat_number]:
                                    sorted_dcm_dict[acq_date][series_time][repeat_number][image_type] = {}
                                    #1 slice
                                    sorted_dcm_dict[acq_date][series_time][repeat_number][image_type][1] = {}
                                sorted_dcm_dict[acq_date][series_time][repeat_number][image_type][1] = sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number][image_number]
                        elif len(sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number]) == 2*n_slices*n_elements:
                            #number of del rec images is elements*slices*2   
                            #individual_elements_image
                            image_type = "DelRec"
                            
                            for image_number in sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number]:
                                
                                element_number = image_number % n_elements
                                if element_number == 0:
                                    element_number = n_elements
                                if element_number == image_number:
                                    repeat_number = 1
                                else:
                                    repeat_number = 2
                                if repeat_number not in sorted_dcm_dict[acq_date][series_time]:
                                    sorted_dcm_dict[acq_date][series_time][repeat_number] = {}
                                if image_type not in sorted_dcm_dict[acq_date][series_time][repeat_number]:
                                    sorted_dcm_dict[acq_date][series_time][repeat_number][image_type] = {}
                                    #1 slice
                                    sorted_dcm_dict[acq_date][series_time][repeat_number][image_type][1] = {}
                                if element_number not in sorted_dcm_dict[acq_date][series_time][repeat_number][image_type][1]:
                                    sorted_dcm_dict[acq_date][series_time][repeat_number][image_type][1][element_number] = {}
                                sorted_dcm_dict[acq_date][series_time][repeat_number][image_type][1][element_number] = sorted_dcm_dict[acq_date][series_time]["unsorted"][series_number][image_number]
                   
                    del sorted_dcm_dict[acq_date][series_time]["unsorted"]            
            return sorted_dcm_dict
     
        
        def Initialise_Directory(self, path):
            """
            Checks if the path exists. If it doesn't create it.
            
            Parameters
            ----------
            path : str
                Should be a directory
            """
            if not os.path.exists(path):
                os.makedirs(path)
        
        def Copy_File(self, current_path, target_directory, new_name):
            """
            Checks if a file already exists in the new location.  If it 
            doesn't the file is copied to the new location and given a new 
            name.
            
            Parameters
            ----------
            current_path : str
                Current location of file
            target_directory : str
                Directory to copy the file too.
            new_name : str
                New name of the file

            """
            self.new_path = os.path.join(target_directory, new_name)                                    
            if not os.path.exists(self.new_path):
                self.Initialise_Directory(target_directory)
                shutil.copyfile(current_path, self.new_path)
        
        def Archive_Images(self, dcm_dict, archive_root):
            """
            Parameters
            ----------
            dcm_dict : dict
                Dict of the form:
                    {<Date>:{<Series Time>:
                {<repeat_number>:{"DelRec":{<Slice number>:{<element number>:
                                        {"path":<path>, "dcm":<DICOM>}}},
                                "Combined":{<Slice number>:
                                        {"path":<path>, "dcm":<DICOM>}}}}}}
            archive_root : str
                base directory to archive images to 
                
            archive DICOMs in a sorted format.
            """
            
            for acq_date in dcm_dict:
                for series_time in dcm_dict[acq_date]:
                    for acq_time in dcm_dict[acq_date][series_time]:
                        for img_type in dcm_dict[acq_date][series_time][acq_time]:
                            
                            for slice_n in dcm_dict[acq_date][series_time][acq_time][img_type]:
                                if img_type == "DelRec":
                                    for element_n in dcm_dict[acq_date][series_time][acq_time][img_type][slice_n]:
                                        current_path = dcm_dict[acq_date][series_time][acq_time][img_type][slice_n][element_n]["path"]
                                        target_directory =  os.path.join(archive_root, acq_date, img_type)
                                        new_name =str(series_time) + "_" + str(acq_time) + "_" + img_type + "_" + str(slice_n) + "_" + str(element_n)
                                        self.Copy_File(current_path, target_directory, new_name)
                                else:
                                    try:
                                        current_path = dcm_dict[acq_date][series_time][acq_time][img_type][slice_n]["path"]
                                        target_directory =  os.path.join(archive_root, acq_date, img_type)
                                        new_name =str(series_time) + "_" + str(acq_time) +  "_" + img_type + "_" + str(slice_n)
                                        self.Copy_File(current_path, target_directory, new_name)
                                    except KeyError:
                                        pass
                                
        
        def Convert_to_PNG(self, dcm_array, target_directory, new_name):
            """
            Parameters
            ----------
            dcm_array : np array
                Dicom.pixel_array to produce PNG image from
            target_directory : str
                Directory to save image in
            new_name : str
                Filename to save PNG as   
            Saves the DICOM Pixel array as a PNG
            
            Convert DICOM.PixelArray to a PNG and save it in the specified directory
            """
            #Convert to float to avoid overflow or underflow losses.
            image_2d = dcm_array
        
            # Rescaling grey scale between 0-255
            image_2d_scaled = (np.maximum(image_2d,0) / image_2d.max()) * 255.0
        
            # Convert to uint
            image_2d_scaled = np.uint8(image_2d_scaled)
            
            cv2.imwrite(os.path.join(target_directory, new_name),image_2d_scaled)
                                
                 
        
        def Archive_PNGs(self, dcm_dict):
            """
            Parameters
            ----------
            dcm_dict : dict
                Dict of the form:
                    {<Date>:{<Series Time>:
                {<repeat_number>:{"DelRec":{<Slice number>:{<element number>:
                                        {"path":<path>, "dcm":<DICOM>}}},
                                "Combined":{<Slice number>:
                                        {"path":<path>, "dcm":<DICOM>}}}}}}
            loops through dcm_dict and saves PNG images from each of the DICOM.Pixelelements
            
            """            
            #root path to save PNG images
            coil_root = os.path.join(self.png_archive, self.scanner_name, self.coil_name)
            
            for acq_date in dcm_dict:
                for series_time in dcm_dict[acq_date]:
                    for acq_time in dcm_dict[acq_date][series_time]:
                        for img_type in dcm_dict[acq_date][series_time][acq_time]:
                            #directory to save PNG images in
                            target_directory =  os.path.join(coil_root, acq_date, img_type)
                            self.Initialise_Directory(target_directory)         
                            for slice_n in dcm_dict[acq_date][series_time][acq_time][img_type]:
                                if img_type == "DelRec":
                                    for element_n in dcm_dict[acq_date][series_time][acq_time][img_type][slice_n]:
                                        dcm = dcm_dict[acq_date][series_time][acq_time][img_type][slice_n][element_n]["dcm"]
                                        #file name of png to be produced
                                        dcm_array = dcm.pixel_array.astype(float)
                                        new_name = "slice_" + str(slice_n) + "_element_" + str(element_n) + ".png"
                                        self.Convert_to_PNG(dcm_array, target_directory, new_name)
                                         
                                else:
                                    try:
                                        dcm = dcm_dict[acq_date][series_time][acq_time][img_type][slice_n]["dcm"]
                                        dcm_array = dcm.pixel_array.astype(float)
                                    except KeyError:
                                        dcm_array = dcm_dict[acq_date][series_time][acq_time][img_type][slice_n]["pixel_array"]
                                    #file name of png to be produced
                                    new_name = "slice_" + str(slice_n) + ".png"
                                    self.Convert_to_PNG(dcm_array, target_directory, new_name)
                        
        
        
        
        def Ask_Archive(self):
            self.archive = messagebox.askyesno("Archive DICOMs?")   
        
        def Ask_Figures(self):
            self.figures = messagebox.askyesno("Produce PNGs")   
        
        def Export_To_Excel(self):
            """
            Ask user to select whether to export to excel or not.  Initialises 
            self.excel_export.
            """
            self.excel_export = messagebox.askyesno("Export to Excel?")   
    
            
    class initialise_masks:
        """
        Assumption made that all images selected are in the same location
        Masks are initially produced using the thresholds established in the class initialise_analysis.
        The visualised mask should be created using the combined element image
        and the ideal mask should cover the entire phantom.
        To produce the phantom mask the accepted mask is contracted by 4 pixels and to 
        produce the in air mask the accepted mask is inverted and contracted by 5 pixels.
        
        A mask can be established for multiple slice images however this is only included as a legacy: ideally single slice images should used.
        """
        #low pass filter used to smooth  out noise in image
        low_pass_filter = np.array([[1,2,1],[2,4,2],[1,2,1]])/16
        def __init__(self, dcm_dict, lower_threshold=0.1, upper_threshold=0.1):
            self.mask_dict = {}
            self.lower_threshold = lower_threshold
            self.upper_threshold = upper_threshold
            
            #combined image required for generating mask: loop through dictionary to find one
            for acq_date in dcm_dict:
                for series_time in dcm_dict[acq_date]:
                    for acq_time in dcm_dict[acq_date][series_time]:
                        for slice_n in dcm_dict[acq_date][series_time][acq_time]["Combined"]:
                            if slice_n not in self.mask_dict.keys():
                                self.mask_dict[slice_n] = {}
                            if self.mask_dict[slice_n]=={}:
                                try:
                                    #image to establish mask from
                                    dcm = dcm_dict[acq_date][series_time][acq_time]["Combined"][slice_n]["dcm"]
                                    #image mask is the mask visualised, phantom mask and air mask are used for analysis
                                    dcm_array = dcm.pixel_array
                                    phantom_mask, air_mask, img_mask = self.Get_Masks(dcm_array)
                                    self.mask_dict[slice_n]["phantom"] = phantom_mask
                                    self.mask_dict[slice_n]["air"] = air_mask
                                except KeyError:
                                    try:
                                        dcm_array = dcm_dict[acq_date][series_time][acq_time]["Combined"][slice_n]["pixel_array"]
                                        phantom_mask, air_mask, img_mask = self.Get_Masks(dcm_array)
                                        self.mask_dict[slice_n]["phantom"] = phantom_mask
                                        self.mask_dict[slice_n]["air"] = air_mask
                                    except KeyError:
                                        pass
                                    
                                
                                
                                
                
        def Fig2Img(self, fig):
            """Convert a Matplotlib figure to a PIL Image and return it"""
            import io
            buf = io.BytesIO()
            fig.savefig(buf)
            buf.seek(0)
            img = Image.open(buf)
            return img
     
            
        def Display_Mask(self, img):
            """
            Parameters
            ----------
            img : PIL Image array
                Image to be displayed in tkinter window
            Return True/ False
            
            Displays img and asks the user to confirm if mask is acceptable.  
            If the initial mask isn't acceptable the thresholds should be 
            adjusted however using different thresholds may indicate acoil failure
            """    
            global img_tk
            top = tk.Toplevel()
            top.title("Is the mask acceptable?")
            top.geometry('800x700')
            top.wm_attributes('-topmost', 1)
            
            #adjust size of image so displays in tkinter label
            scale_factor = 800/img.size[0]
            size = tuple((np.array(img.size) * scale_factor).astype(int))
            img_rs = img.resize(size, resample=Image.NEAREST)
            img_tk = ImageTk.PhotoImage(img_rs)
            tk.Label(top, image=img_tk).grid(row=0, column=0, columnspan = 2)
            proceed = tk.IntVar()
            
            #initialise thresholds
            lower_threshold = tk.DoubleVar()
            upper_threshold = tk.DoubleVar()
            lower_threshold.set(self.lower_threshold)
            upper_threshold.set(self.upper_threshold)
            
            tk.Button(top, text="Yes", command=lambda:  proceed.set(True)).grid(row=1, column=0)
            tk.Button(top, text="Retry", command=lambda:  proceed.set(False)).grid(row=1, column=1)
            tk.Label(top, text = "Lower_threshold").grid(row=2, column=0)
            tk.Entry(top, width = 15, textvariable= lower_threshold).grid(row=2, column=1)
            tk.Label(top, text = "Upper_threshold").grid(row=3, column=0)
            tk.Entry(top, width = 15, textvariable= upper_threshold).grid(row=3, column=1)        
            
            #Wait until the user has decided if the mask is acceptable
            top.wait_variable(proceed)
            top.destroy()
            
            #replace thresholds with those specified by the user
            self.lower_threshold = lower_threshold.get()
            self.upper_threshold = upper_threshold.get()
            return proceed.get()
        
        def Expand_and_Contract_Masks(self, original_mask, x_size, y_size, expand=True):
            """Expand/contract original_mask by 1 pixel in all directions;
            x_size and y_size are the dimensions of the image.
            """
            new_mask = np.full_like(original_mask,False)
            for x in range(x_size):
                for y in range(y_size):
                    if original_mask[x,y] == True:
                        if np.all(original_mask[x-1:x+2,y-1:y+2]):
                            new_mask[x,y] = True
                        else:
                            if expand == True:
                                for i in range(3):
                                    for j in range(3):
                                        new_mask[x-1+i,y-1+j] = True
                            else:
                                new_mask[x,y] = False
            return new_mask
            
        
        def Get_Masks(self,dcm_array):
            """
            

            Parameters
            ----------
            dcm : DICOM
                Combined array DICOM to create mask from

            Returns
            -------
            phantom_mask_npbool : NP Array of Booleans
                img_mask_npbool contracted by 4 pixels. Phantom mask ignoring edge effects   
            air_mask_npbool : NP Array of Booleans
                img_mask_npbool inverted and contracted by 5 pixels. Air mask ignoring edge effects 
            img_mask_npbool : NP Array of Booleans
                Accepted mask covering entirity of phantom
            
            Produces a mask based on previously defined thresholds. 
            Asks the user if the mask is acceptable.  If it isn't the user can 
            adjust the thresholds until the mask is acceptable.  It should be 
            noted if theresholds need adjusting this may indicate a coil failure.
            
            The accepted mask is contracted by 4 pixels to produce a mask of 
            the phantom ignoring edge effects.  The accepted mask is also inverted and 
            econtracted by 5 pixels to produce a mask of the air.
            """
            x_size = dcm_array.shape[0]
            y_size = dcm_array.shape[1]
            max_int = np.amax(dcm_array)
            mask_acceptable = False
            #smooth out noise in the image
            dcm_filtered = cv2.filter2D(dcm_array,-1,self.low_pass_filter)
            
            #Loop until User accepts the mask
            while mask_acceptable ==False:
                markers = np.zeros(dcm_filtered.shape)
                #probably air
                markers[dcm_filtered<max_int*self.lower_threshold]=1
                #probably phantom
                markers[dcm_filtered>max_int*self.upper_threshold]=2
                
                #perform watershed algorithm based on the thresholded image 
                #and the intensities of the voxels in the smoothed DICOM
                labels = seg.watershed(dcm_filtered, markers)
                img_mask = np.full_like(dcm_filtered, False)
                img_mask[labels==2]=True
                
                test = np.ma.masked_where(img_mask==False, img_mask)
                fig, ax = plt.subplots()
                #create figure of mask overlaying the DICOM PixelArray
                ax.imshow(dcm_filtered, cmap='gray', interpolation='none', alpha=0.4)
                ax.imshow(test, 'spring', interpolation='none', alpha=0.2)
                PIL_img = self.Fig2Img(fig)
                #close plot: will be displayed in tkinter window
                plt.close(fig)
                #Display figure in tkinter window and ask user if the mask is OK 
                if self.Display_Mask(PIL_img) == True:
                    mask_acceptable = True
            
            #contract accepted mask by 4 pixels to produce phantom_mask
            phantom_mask = img_mask
            for contract_px in range(4):
                contracted_mask = self.Expand_and_Contract_Masks(phantom_mask, x_size, y_size, expand=False)
                phantom_mask = contracted_mask
            
            
            #Invert accepted mask and contract by 5 pixels to produce air_mask
            air_mask = np.full_like(dcm_filtered, True)
            air_mask[img_mask ==True]=False
            for contract_px in range(5):
                contracted_mask = self.Expand_and_Contract_Masks(air_mask, x_size, y_size, expand=False)
                air_mask = contracted_mask
            
            phantom_mask_npbool = phantom_mask.astype(np.bool_)
            air_mask_npbool = air_mask.astype(np.bool_)
            img_mask_npbool = img_mask.astype(np.bool_)
            
            
            return phantom_mask_npbool, air_mask_npbool, img_mask_npbool
        
    class calculate_results:
        """Calculate SNR via 3 methods for individual element and combined images:
            -NEMA: SNR calculated via subtraction method; requires 2 images
            -noise_std: SNR calculated using the assumption that the noise in 
            the image is proportional to the S.D. of the signal in air.
            -noise_av (preferred mathod): SNR calculated using the assumption 
            that the noise in the image is proportional to the mean of the 
            signal in air.
        The uniformity of the combined element image is also calculated via 
        the method outlined in IPEM Report 112 and is refereed to as the 
        integral uniformity method. A low pass filter is applied to the image 
        and the "phantom signal mask" (used for SNR calculation) is applied 
        to the image so only voxels only within the phantom are considered. 
        """
        #low pass filter used for uniformity calculation
        low_pass_filter = np.array([[1,2,1],[2,4,2],[1,2,1]])/16
        #constant of proportionality for noise via "noise_std"/"noise_av" 
        #methods depends on the number of elements the image is produced from
        element_scale_factors = {1:{"SD":0.6551,"Mean":1.2533},
                               2:{"SD":0.6824,"Mean":1.8800},
                               3:{"SD":0.6911,"Mean":2.3500},
                               4:{"SD":0.6953,"Mean":2.7416},
                               6:{"SD":0.6994,"Mean":3.3928},
                               8:{"SD":0.7014,"Mean":3.9380},
                               12:{"SD":0.7035,"Mean":4.8271},
                               14:{"SD":0.7040,"Mean":5.2243},
                               15:{"SD":0.7042,"Mean":5.4123},
                               16:{"SD":0.7043,"Mean":5.6128},
                               18:{"SD":0.7046,"Mean":5.9585},
                               20:{"SD":0.7049,"Mean":6.2698},
                               28:{"SD":0.7055,"Mean":7.4425},
                               30:{"SD":0.7056,"Mean":7.7083},
                               32:{"SD":0.7057,"Mean":7.9688},
                               64:{"SD":0.7064,"Mean":11.2916}
                               }
        def __init__(self, dcm_dict, mask_dict, n_elements):
            #create dictionary to store data in
            self.Initialise_SNR_Dict(dcm_dict)
            #loop through the dicoms in dcm_dict calculating the SNR for them 
            #all and the uniformity of the combined element images, store the 
            #data in self.SNR_Dict
            self.Loop_Dicoms(dcm_dict, mask_dict, n_elements)
            
        
        def Initialise_SNR_Dict(self, dcm_dict):
            """
            Parameters
            ----------
            dcm_dict : sorted dictionary of DICOMS
            
            Initialises self.SNR_Dict: a dictionary to store the results for all the dicoms inside scm_dict

            """
            #SNR calculated via 3 methods
            snr_types = ["NEMA", "noise_std", "noise_av"]
            self.SNR_Dict = {}
            for snr_type in snr_types:
                self.SNR_Dict[snr_type] = {}
                for date in dcm_dict:
                    self.SNR_Dict[snr_type][date] = {}
                    for series_time in dcm_dict[date]:
                        for acquisition_time in dcm_dict[date][series_time]:
                            acq_ID = str(series_time) + str(acquisition_time)
                            self.SNR_Dict[snr_type][date][acq_ID] ={}
                            self.SNR_Dict[snr_type][date][acq_ID]["Combined"] = {}
                            if "DelRec" in dcm_dict[date][series_time][acquisition_time]:
                                self.SNR_Dict[snr_type][date][acq_ID]["DelRec"] = {}
                                for slice_n in dcm_dict[date][series_time][acquisition_time]["Combined"]:
                                    self.SNR_Dict[snr_type][date][acq_ID]["Combined"][slice_n] = {}
                                    self.SNR_Dict[snr_type][date][acq_ID]["DelRec"][slice_n] = {}
                                    for element_n in dcm_dict[date][series_time][acquisition_time]["DelRec"][slice_n]:
                                        self.SNR_Dict[snr_type][date][acq_ID]["DelRec"][slice_n][element_n] = {}
                            else:
                                for slice_n in dcm_dict[date][series_time][acquisition_time]["Combined"]:
                                    self.SNR_Dict[snr_type][date][acq_ID]["Combined"][slice_n] = {}
        
        def Sort_Results(self, SNR_results):
            """
            Sort SNR_results into format ready for export to excel.
            4 excel files are produced for each type of SNR calculation:
                -del_rec: all SNR results from the individual elements
                -del_rec_group:  For each array an average of all the results 
                    from this analysis run
                -combined: all SNR and uniformity results from the combined elements
                -combined_group:  For combined image an average of all the results 
                    from this analysis run
            """
            self.sorted_SNR_results = {}
            
            #Initialis headings.  These headings are the become the first row of the Excel files.
            self.combined_base_headings_list = []
            self.delrec_base_headings_list = []
            self.combined_group_headings_list = []
            self.delrec_group_headings_list = []
            for SNR_type in SNR_results:
                if SNR_type not in self.sorted_SNR_results:
                    self.sorted_SNR_results[SNR_type] = {}
                for date in SNR_results[SNR_type]:
                    for acq_time in SNR_results[SNR_type][date]:
                        for image_type in SNR_results[SNR_type][date][acq_time]:
                            if image_type not in self.sorted_SNR_results[SNR_type]:
                                self.sorted_SNR_results[SNR_type][image_type] = {}
                            if image_type == "Combined":
                                if date not in self.sorted_SNR_results[SNR_type][image_type]:
                                    self.sorted_SNR_results[SNR_type][image_type][date] = {}
                                if acq_time not in self.sorted_SNR_results[SNR_type][image_type][date]:
                                    self.sorted_SNR_results[SNR_type][image_type][date][acq_time] = {}
                                results = {"Acquisition_Date": date, "Acquisition_Time": acq_time}
                                for slice_n in SNR_results[SNR_type][date][acq_time][image_type]:
                                    results.update(SNR_results[SNR_type][date][acq_time][image_type][slice_n])
                                self.sorted_SNR_results[SNR_type][image_type][date][acq_time] = results
                                if self.combined_base_headings_list == []:
                                    self.combined_base_headings_list = list(self.sorted_SNR_results[SNR_type][image_type][date][acq_time].keys())
                                    
                            else:
                                for slice_n in SNR_results[SNR_type][date][acq_time][image_type]:
                                    if slice_n not in self.sorted_SNR_results[SNR_type][image_type]:
                                        self.sorted_SNR_results[SNR_type][image_type][slice_n] = {}
                                    if date not in self.sorted_SNR_results[SNR_type][image_type][slice_n]:
                                        self.sorted_SNR_results[SNR_type][image_type][slice_n][date] = {}
                                    if acq_time not in self.sorted_SNR_results[SNR_type][image_type][slice_n][date]:
                                        self.sorted_SNR_results[SNR_type][image_type][slice_n][date][acq_time] = {}
                                    results = {"Acquisition_Date": date, "Acquisition_Time": acq_time}
                                    for element_n in SNR_results[SNR_type][date][acq_time][image_type][slice_n]:
                                        results.update(SNR_results[SNR_type][date][acq_time][image_type][slice_n][element_n])
                                    self.sorted_SNR_results[SNR_type][image_type][slice_n][date][acq_time]= results
                                if self.delrec_base_headings_list == []:
                                    self.delrec_base_headings_list = list(self.sorted_SNR_results[SNR_type][image_type][slice_n][date][acq_time].keys())
        
            for heading in self.combined_base_headings_list:
                if heading in ["Acquisition_Date", "Acquisition_Time"]:
                    self.combined_group_headings_list.append(heading)
                elif "Group" in heading:
                    self.combined_group_headings_list.append(heading)
            for heading in self.delrec_base_headings_list:
                if heading in ["Acquisition_Date", "Acquisition_Time"]:
                    self.delrec_group_headings_list.append(heading)
                elif "Group" in heading:
                    self.delrec_group_headings_list.append(heading)
        
        
        
        def SNR_Moriel(self,img_arr,phantom_mask,air_mask,n_elements):
            """
            Parameters
            ----------
            img_arr : DICOM.PixelArray
                image to calculate SNR of
            phantom_mask : np Boolean array
                mask of phantom ignoring edge effects
            air_mask : np Boolean array
                mask of air ignoring edge effects
            n_elements : int
                Number of elements used to generate img_arr: SNR scale factors 
                depending on number of elements

            Returns
            -------
            SNR_MORIEL_std : Float
                SNR calculated via noise_std method
            SNR_MORIEL_av : Float
                SNR calculated via noise_av method

            Calulates SNR via two methods using a simgle image:
                -noise_std: SNR calculated using the assumption that the noise in 
                the image is proportional to the S.D. of the signal in air.
                -noise_av (preferred mathod): SNR calculated using the assumption 
                that the noise in the image is proportional to the mean of the 
                signal in air.
            """
            #mask img_arr: only phantom
            signal_arr = img_arr[phantom_mask]
            #mask img_arr: only air
            noise_arr = img_arr[air_mask]
            #get scale factors for SNR calculations based on the number of elements
            scale_factors = self.element_scale_factors[n_elements]
            
            signal_av = np.mean(signal_arr)
            
            #scaled noise
            noise_std = np.std(noise_arr)/scale_factors["SD"]
            noise_av = np.mean(noise_arr)/scale_factors["Mean"]
            #calculate SNR
            SNR_MORIEL_std = round(self.bandwidth_scalar*signal_av/noise_std,2)
            SNR_MORIEL_av = round(self.bandwidth_scalar*signal_av/noise_av,2)
            return SNR_MORIEL_std, SNR_MORIEL_av
            
        def SNR_NEMA(self,img_arr_1,img_arr_2,phantom_mask):
            """
            Calculates SNR via NEMA "subtraction" method: requires two 
            identically acquired images (img_arr_1 and img_arr_2) and a 
            mask of the phantom

            """
            signal_arr= img_arr_1[phantom_mask]
            noise_arr = np.int16(np.subtract(np.int32(img_arr_1), np.int32(img_arr_2)))[phantom_mask]
            signal_av = np.mean(signal_arr)
            noise_std = np.std(noise_arr)
            SNR_NEMA = round(self.bandwidth_scalar*(2**0.5)*signal_av/noise_std,2)
            return SNR_NEMA
            
        def single_img_SNR(self,signal_array,phantom_mask,air_mask,n_elements):
            """Calulates SNR via two methods using a simgle image and put 
            results in self.SNR_results"""
            SNR_std, SNR_av = self.SNR_Moriel(signal_array,phantom_mask,air_mask,n_elements)
            self.SNR_results["noise_std"].append(SNR_std)
            self.SNR_results["noise_av"].append(SNR_av)
         
        def SNR_Calculate_dcm(self, dcm1, phantom_mask, air_mask, dcm2=None, n_elements=1):
            """Calulates SNR via three methods using dcm1 and dmc2.  One 
            result is produced via the NEMA method which requires 2 images and 
            two results are produced via the the noise_std and noise_av: one for
            each DICOM."""
            self.SNR_results = {}
            self.SNR_results["noise_std"] = []
            self.SNR_results["noise_av"] = []
            #Images should all be acquired at a bandwidth of 222 Hz/Px.  The 
            #scalar allows images which are acqhuired at a different bandwidth to still be used
            try:
                self.bandwidth_scalar =  ((dcm1.PixelBandwidth/130)**0.5)
            except AttributeError:
                #Some Dicom headers don't possess the atribute PixelBandwidth 
                #(the bandwidth is still in the header but pydicom cant access 
                #via DICOM.PixelBandwidth).  The BW is assumed to be 222 Hz/px
                self.bandwidth_scalar =  ((222/130)**0.5)
            
            signal_array_1 = dcm1.pixel_array
            self.single_img_SNR(signal_array_1, phantom_mask, air_mask, n_elements)
            try:
                signal_array_2 = dcm2.pixel_array
                self.single_img_SNR(signal_array_2, phantom_mask, air_mask, n_elements)
                self.SNR_results["NEMA"] = self.SNR_NEMA(signal_array_1,signal_array_2,phantom_mask)
            except  IndexError:
                #CHECK CORRECT error
                pass
            return self.SNR_results    
        
        def SNR_Calculate_array(self, signal_array_1, phantom_mask, air_mask, signal_array_2=None, n_elements=1):
            """Calulates SNR via three methods using dcm1 and dmc2.  One 
            result is produced via the NEMA method which requires 2 images and 
            two results are produced via the the noise_std and noise_av: one for
            each DICOM."""
            self.SNR_results = {}
            self.SNR_results["noise_std"] = []
            self.SNR_results["noise_av"] = []
            self.bandwidth_scalar =  ((222/130)**0.5)
            
            
            self.single_img_SNR(signal_array_1, phantom_mask, air_mask, n_elements)
            try:
                self.single_img_SNR(signal_array_2, phantom_mask, air_mask, n_elements)
                self.SNR_results["NEMA"] = self.SNR_NEMA(signal_array_1,signal_array_2,phantom_mask)
            except  IndexError:
                #CHECK CORRECT error
                pass
            return self.SNR_results
        
        def Uniformity_Calculate(self, dcm_array, phantom_mask):
            """Calculate the uniformity of dcm via 
            the method outlined in IPEM Report 112 and is refereed to as the 
            integral uniformity method. A low pass filter is applied to the image 
            and the "phantom signal mask" (used for SNR calculation) is applied 
            to the image so only voxels only within the phantom are considered. 
            """
            dcm_filtered = cv2.filter2D(dcm_array,-1,self.low_pass_filter)
            ROI_voxels = dcm_filtered[phantom_mask]
            max_int = np.max(ROI_voxels)
            min_int = np.min(ROI_voxels)
            uniformity = 1-((max_int-min_int)/(max_int+min_int))
            return uniformity
        
        def Loop_Dicoms(self, dcm_dict, mask_dict, n_elements):
            """
            Loop through dcm_dict calculating the SNR for all images via 3 methods 
            (NEMA,noise_std and noise_av) and the uniformity of the combined element images
            
            Also calculates group averages for this analysis run
            
            Sorts results ready for export to excel
            """
            for date in dcm_dict:
                for series_time in dcm_dict[date]:
                    acq_times = list(dcm_dict[date][series_time].keys())
                    acq_ID_1 = str(series_time) + str(acq_times[0])
                    acq_ID_2 = str(series_time) + str(acq_times[1])
                       
                    del self.SNR_Dict["NEMA"][date][str(series_time) + str(acq_times[1])]
                    for img_type in dcm_dict[date][series_time][acq_times[0]]:
                        for slice_n in dcm_dict[date][series_time][acq_times[0]][img_type]:
                            if img_type == "Combined":
                                try:
                                    dcm_1 = dcm_dict[date][series_time][acq_times[0]][img_type][slice_n]["dcm"]
                                    dcm_2 = dcm_dict[date][series_time][acq_times[1]][img_type][slice_n]["dcm"]
                                    SNRs = self.SNR_Calculate_dcm(dcm_1, mask_dict[slice_n]["phantom"], mask_dict[slice_n]["air"], dcm_2, n_elements)
                                    dcm_array_1 = dcm_1.pixel_array
                                    dcm_array_2 = dcm_2.pixel_array
                                except KeyError:
                                    dcm_array_1 = dcm_dict[date][series_time][acq_times[0]][img_type][slice_n]["pixel_array"]
                                    dcm_array_2 = dcm_dict[date][series_time][acq_times[1]][img_type][slice_n]["pixel_array"]
                                    SNRs = self.SNR_Calculate_array(dcm_array_1, mask_dict[slice_n]["phantom"], mask_dict[slice_n]["air"], dcm_array_2, n_elements)
                                uniformity_1 = round(self.Uniformity_Calculate(dcm_array_1, mask_dict[slice_n]["phantom"]),2)
                                uniformity_2 = round(self.Uniformity_Calculate(dcm_array_2, mask_dict[slice_n]["phantom"]),2)
                                self.SNR_Dict["NEMA"][date][acq_ID_1]["Combined"][slice_n] = {"SNR_slice_{}".format(slice_n):SNRs["NEMA"],"Uniformity_slice_{}".format(slice_n):uniformity_1}
                                self.SNR_Dict["noise_std"][date][acq_ID_1]["Combined"][slice_n] = {"SNR_slice_{}".format(slice_n):SNRs["noise_std"][0],"Uniformity_slice_{}".format(slice_n):uniformity_1}
                                self.SNR_Dict["noise_std"][date][acq_ID_2]["Combined"][slice_n] = {"SNR_slice_{}".format(slice_n):SNRs["noise_std"][1],"Uniformity_slice_{}".format(slice_n):uniformity_2}
                                self.SNR_Dict["noise_av"][date][acq_ID_1]["Combined"][slice_n] = {"SNR_slice_{}".format(slice_n):SNRs["noise_av"][0],"Uniformity_slice_{}".format(slice_n):uniformity_1}
                                self.SNR_Dict["noise_av"][date][acq_ID_2]["Combined"][slice_n] = {"SNR_slice_{}".format(slice_n):SNRs["noise_av"][1],"Uniformity_slice_{}".format(slice_n):uniformity_2}
                            else:
                                #delrec image
                                for element_n in dcm_dict[date][series_time][acq_times[0]][img_type][slice_n]:
                                    dcm_1 = dcm_dict[date][series_time][acq_times[0]][img_type][slice_n][element_n]["dcm"]
                                    dcm_2 = dcm_dict[date][series_time][acq_times[1]][img_type][slice_n][element_n]["dcm"]
                                    SNRs = self.SNR_Calculate_dcm(dcm_1, mask_dict[slice_n]["phantom"], mask_dict[slice_n]["air"], dcm_2, 1)
                                    self.SNR_Dict["NEMA"][date][acq_ID_1]["DelRec"][slice_n][element_n] = {"SNR_element_{}".format(element_n):SNRs["NEMA"]}
                                    self.SNR_Dict["noise_std"][date][acq_ID_1]["DelRec"][slice_n][element_n] = {"SNR_element_{}".format(element_n):SNRs["noise_std"][0]}
                                    self.SNR_Dict["noise_std"][date][acq_ID_2]["DelRec"][slice_n][element_n] = {"SNR_element_{}".format(element_n):SNRs["noise_std"][1]}
                                    self.SNR_Dict["noise_av"][date][acq_ID_1]["DelRec"][slice_n][element_n] = {"SNR_element_{}".format(element_n):SNRs["noise_av"][0]}
                                    self.SNR_Dict["noise_av"][date][acq_ID_2]["DelRec"][slice_n][element_n] = {"SNR_element_{}".format(element_n):SNRs["noise_av"][1]}
            
            
                      
            #calculate the group averages for this analysis run
            for SNR_type in self.SNR_Dict:
                for date in self.SNR_Dict[SNR_type]:
                        acq_IDs = list(self.SNR_Dict[SNR_type][date].keys())
                        if "DelRec" in self.SNR_Dict[SNR_type][date][acq_IDs[0]]:
                            slices = list(self.SNR_Dict[SNR_type][date][acq_IDs[0]]["DelRec"].keys())
                            elements = list(self.SNR_Dict[SNR_type][date][acq_IDs[0]]["DelRec"][slices[0]].keys())
                            img_types = ["DelRec", "Combined"]
                        else:
                            slices = list(self.SNR_Dict[SNR_type][date][acq_IDs[0]]["Combined"].keys())
                            img_types = ["Combined"]                        
                        for img_type in img_types:
                            for slice_n in slices:
                                if img_type == "DelRec":
                                    for element_n in elements:
                                        SNRs = []
                                        for acq_ID in acq_IDs:
                                            SNRs.append(self.SNR_Dict[SNR_type][date][acq_ID][img_type][slice_n][element_n]["SNR_element_{}".format(element_n)])
                                        mean_SNR = sum(SNRs)/len(SNRs)
                                        group_SNRs = {"Group_SNR_element_{}".format(element_n):round(mean_SNR,2)}
                                        for acq_ID in acq_IDs:
                                            self.SNR_Dict[SNR_type][date][acq_ID][img_type][slice_n][element_n].update(group_SNRs)
                                else:
                                    SNRs = []
                                    uniformities  = []
                                    for acq_ID in acq_IDs:
                                        SNRs.append(self.SNR_Dict[SNR_type][date][acq_ID][img_type][slice_n]["SNR_slice_{}".format(slice_n)])
                                        uniformities.append(self.SNR_Dict[SNR_type][date][acq_ID][img_type][slice_n]["Uniformity_slice_{}".format(slice_n)])
                                    mean_SNR = sum(SNRs)/float(len(SNRs))
                                    mean_uniformity = sum(uniformities)/float(len(uniformities))
                                    group_SNRs = {"Group_SNR_slice_{}".format(slice_n):round(mean_SNR,2), "Group_uniformity_slice_{}".format(slice_n):round(mean_uniformity,2)}
                                    for acq_ID in acq_IDs:
                                        self.SNR_Dict[SNR_type][date][acq_ID][img_type][slice_n].update(group_SNRs)    
                            
                
            self.Sort_Results(self.SNR_Dict)     
  
      

            
    class export_to_excel:
        """
            Export results to excel.  The preferred SNR analysis method is noise_av.
            4 excel files are produced for each of the three SNR calculations (12 in total):
                -del_rec: all SNR results from the individual elements
                -del_rec_group:  For each element an average of all the results 
                    from this analysis run
                -combined: all SNR and uniformity results from the combined elements
                -combined_group:  For combined image an average of all the results 
        """          
        results_base_directory = "S:/Radiology-Directorate/IRIS/Non-Ionising/MRI/QA and Calibration/Quarterly QC/SNR QC/Results_Sheets/Moriel"
        scanner_name_dict = {"45504":"PHT_ElitionX_MR3", "183109": "DCH_Sola_MR2", "202541": "RBH_Sola_MR2", "183188": "RBH_Sola_MR3", "169696": "DCH_Avanto_MR1", "64114": "RBH_Essenza_MR3", "169641": "RBH_Avanto_MR1", "41113": "RBH_Aera_MR2", "47009": "PHT_Ambition_X", "169672":"PHT_Avanto", "42044": "PHT_Ingenia3T_MR3", "2016": "AECC_Paramed_MROPEN"}        
        
        def __init__(self, SNRs, scanner_ID, coil_name):
            """

            Parameters
            ----------
            SNRs : Dictionary of results
            scanner_ID : Scanner ID from DICOM Header
            coil_name : Name of coil used to acquire images
            """
            #base directory containing all results for the scanner specified
            self.results_base_directory_scanner = os.path.join(self.results_base_directory,self.scanner_name_dict[scanner_ID]) 
            #Export Results
            self.Export_Data(SNRs, self.results_base_directory_scanner, self.scanner_name_dict[scanner_ID], coil_name)
            #Ensure all excel files are closed
            self.wb.close()
        
        
        
        def Open_Spreadsheet(self, directory, file_name):
            
            """
            Parameters
            ----------
            directory : str
                directory the spreadsheet is contained in
            file_name : str
                name of the saved file
            
            Initialises <self.wb> a workbook (requires openpyxl). If <file_name> 
            doesn't exist in the directory specified it is created.
            """
            
            if not os.path.exists(directory):
                    os.makedirs(directory)
            os.chdir(directory)
            
            """Open correct spreadsheet. If it doesn't exist create it"""
            if os.path.isfile(os.path.join(directory, file_name)):
                self.wb = openpyxl.load_workbook(file_name)
            else:
                #first SNR data for this scanner; create new file
                self.wb = openpyxl.Workbook()
                self.wb.save(file_name)
                print("New Spreadsheet created.  This should be the first SNR measurement for this scanner")
            
    
        def Initialise_Sheet(self, ws_name, headings):
            """
            Parameters
            ----------
            ws_name : str
                Name of worksheet required for initialisation
            headings: list
            
            Create a work sheet in the workbook <self.wb>, called <ws_name>.  
            The works sheet will be initialised with headings specified by 
            <headings>. Initialised <self.ws> (requires openpyxl).
            """
            self.wb.create_sheet(ws_name)
            self.ws = self.wb[ws_name]
            for i in range(len(headings)):
                #write <headings> in the first row of the worksheet
                self.ws.cell(column=i+1, row=1, value=headings[i])
    
        def Open_Sheet(self, file_name, sheet_name, headings_list= None):
            """
            Parameters
            ----------
            file_name : str
            sheet_name : str
            headings_list : list
            
            Initialise self.ws.  If <sheet_name> doesn't exist in <file_name>, 
            create it with headings specified by <headings_list>.
            
            """
            if not sheet_name in self.wb.sheetnames:
                #first SNR data for this coil; create new file
                if headings_list == None:
                    print("Please specify headings_list")
                else:
                    self.Initialise_Sheet(ws_name = sheet_name, headings = headings_list)
                    self.wb.save(file_name)
                    print("New Sheet created.  This should be the first SNR measurement for this coil")
            else:
                self.ws = self.wb[sheet_name]
    
        def Get_Previous_Analysises_Performed(self):
            """
            Returns
            ---------
            previous_results_details : Dictionary of the form{<date>:[<times>]}
                Containing dates and times of previous analysises in self.ws
            headings : list
                List of headings in self.ws
            
            Looks in the first two columns of <self.ws> and returns two lists of all 
            the elements in these to columns.  These columns contain the acquisition 
            dates and acquisition times of the images on which SNR analysis has 
            previously been performed. 
            
            The first row is assumed to contain the headings
            """
            
            dates_list = []
            times_list = [] 
            headings = []
            previous_results_details = {}
            """loop through rows, column A containes a list of acquisition dates and 
            column B contains a list of acquisiton times"""
            
            for row in self.ws['A']:
                dates_list.append(row.value)
            for row in self.ws['B']:
                times_list.append(row.value)
            for col in self.ws[1]:
                headings.append(col.value)
                
            for i in range(len(dates_list)):
                if dates_list[i] not in previous_results_details:
                    previous_results_details[dates_list[i]] = []
                previous_results_details[dates_list[i]].append(times_list[i])

            return previous_results_details, headings
    
    
    
        def Export_Data(self, SNRs, directory, file_name_base, coil_name):   
            """
            Parameters
            ----------
            SNRs : Output of class calculate_results containing sorted_SNR_Results and headings for the different spreadsheets
            directory : Directory resluts should be saved in
            file_name_base : Base file name of excel files containing scanner name
            coil_name : COil used for acquisition

            Loops through the dictionary of results saving them to the correct 
            spreadsheet.  Results are only added to the spreadsheet if a previous entry for that date doesn't exist.'

            """            
            #get dictionary of results
            SNR_data_dict = SNRs.sorted_SNR_results
            #Different Excel files for each SNR calculation type
            for SNR_type in SNR_data_dict: 
                #Different Excel files for combined and individual elemnt images
                for img_type in  SNR_data_dict[SNR_type]:
                    if img_type == "Combined":
                        #Export all Combined Element Data
                        base_file_name = SNR_type + "_"+ file_name_base + "_combined.xlsx"
                        self.Open_Spreadsheet(directory, base_file_name)
                        self.Open_Sheet(base_file_name, coil_name, SNRs.combined_base_headings_list)
                        #get previous results and check entry doesn't exist
                        previous_results_details, headings = self.Get_Previous_Analysises_Performed()
                        for date in SNR_data_dict[SNR_type][img_type]:
                            for time in SNR_data_dict[SNR_type][img_type][date]:
                                if date in list(previous_results_details.keys()):
                                    if time in previous_results_details[date]:
                                        print("SNR measurement already performed for this dataset, New analysis has not been inputed")
                                else:
                                    results_to_input = SNR_data_dict[SNR_type][img_type][date][time]
                                    new_row = self.ws.max_row+1
                                    for i in range(len(headings)):
                                        self.ws.cell(column=i+1, row=new_row, value=results_to_input[headings[i]])
                                    self.wb.save(base_file_name)    
                        
                        self.wb.close()
                        
                        #Export group average Combined Element Data
                        group_file_name = SNR_type + "_"+ file_name_base + "_group_combined.xlsx"
                        self.Open_Spreadsheet(directory, group_file_name)
                        self.Open_Sheet(group_file_name, coil_name, SNRs.combined_group_headings_list)
                        #get previous results and check entry doesn't exist
                        previous_results_details, headings = self.Get_Previous_Analysises_Performed()
                        for date in SNR_data_dict[SNR_type][img_type]:
                            time = list(SNR_data_dict[SNR_type][img_type][date].keys())[0]
                            if date in list(previous_results_details.keys()):
                                if time in previous_results_details[date]:
                                    print("SNR measurement already performed for this dataset, New analysis has not been inputed")
                            else:
                                results_to_input = SNR_data_dict[SNR_type][img_type][date][time]
                                new_row = self.ws.max_row+1
                                for i in range(len(headings)):
                                    self.ws.cell(column=i+1, row=new_row, value=results_to_input[headings[i]])
                                self.wb.save(group_file_name)
                        self.wb.close()
                    
                    elif img_type == "DelRec":
                        #Export indiviual element Data
                        delrec_base_file_name = SNR_type + "_"+ file_name_base + "_delrec.xlsx"
                        self.Open_Spreadsheet(directory, delrec_base_file_name)
                        for slice_n in SNR_data_dict[SNR_type][img_type]:
                            self.Open_Sheet(delrec_base_file_name, coil_name+"_slice_"+str(slice_n), SNRs.delrec_base_headings_list)
                            #get previous results and check entry doesn't exist
                            previous_results_details, headings = self.Get_Previous_Analysises_Performed()
                            for date in SNR_data_dict[SNR_type][img_type][slice_n]:
                                for time in SNR_data_dict[SNR_type][img_type][slice_n][date]:
                                    if date in list(previous_results_details.keys()):
                                        if time in previous_results_details[date]:
                                            print("SNR measurement already performed for this dataset, New analysis has not been inputed")
                                    else:
                                        results_to_input = SNR_data_dict[SNR_type][img_type][slice_n][date][time]
                                        new_row = self.ws.max_row+1
                                        for i in range(len(headings)):
                                            try:
                                                self.ws.cell(column=i+1, row=new_row, value=results_to_input[headings[i]])
                                            except KeyError:
                                                pass
                                        self.wb.save(delrec_base_file_name)
                            self.wb.close()
                        
                        #Export group average indiviual element Data
                        group_delrec_base_file_name = SNR_type + "_"+ file_name_base + "_group_delrec.xlsx"
                        self.Open_Spreadsheet(directory, group_delrec_base_file_name)
                        for slice_n in SNR_data_dict[SNR_type][img_type]:
                            self.Open_Sheet(group_delrec_base_file_name, coil_name+"_slice_"+str(slice_n), SNRs.delrec_group_headings_list)
                            #get previous results and check entry doesn't exist
                            previous_results_details, headings = self.Get_Previous_Analysises_Performed()
                            for date in SNR_data_dict[SNR_type][img_type][slice_n]:
                                time = list(SNR_data_dict[SNR_type][img_type][slice_n][date].keys())[0]
                                if date in list(previous_results_details.keys()):
                                    if time in previous_results_details[date]:
                                        print("SNR measurement already performed for this dataset, New analysis has not been inputed")
                                else:
                                    results_to_input = SNR_data_dict[SNR_type][img_type][slice_n][date][time]
                                    new_row = self.ws.max_row+1
                                    for i in range(len(headings)):
                                        try:
                                            self.ws.cell(column=i+1, row=new_row, value=results_to_input[headings[i]])
                                        except KeyError:
                                            pass
                                    self.wb.save(group_delrec_base_file_name)
                            self.wb.close()
    
    class produce_figures:
        """creates png figures to put in the QC report:
            -Tables: Table comparing results to previous baselines
            -graphs: graphs of historic results showing parameter progression 
            over time
        Requires baseline spreadsheets to be set up specifying the expected 
        result for each parameter
        """
        results_base_directory = "S:/Radiology-Directorate/IRIS/Non-Ionising/MRI/QA and Calibration/Quarterly QC/SNR QC/Results_Sheets/Moriel"
        png_archive = "S:/Radiology-Directorate/IRIS/Non-Ionising/MRI/QA and Calibration/Quarterly QC/SNR QC/PNG_images/"
        def __init__(self,results,scanner_name,coil_name):
            self.create_tables(results,scanner_name,coil_name)
            self.create_graphs(scanner_name,coil_name)
            
        def create_tables(self,results_dict,scanner_name,coil_name):
            """
            Create table comparing results to previous baselines
            """
            #combined element results
            results_to_print_c = {}
            #individual element results
            results_to_print_a = {}
            #headings in current baseline spreadsheet
            baseline_headings = []
            #heading of interest in baseline spreadsheet (only group headings)
            headings_to_print = []
            #get keys for one of the results.  This is used to initialise the headings
            dates = list(results_dict["Combined"].keys())
            date = dates[0]
            acq_times = list(results_dict["Combined"][date].keys())
            acq_time = acq_times[0]
            
            file_name = "baseline_noise_av_" + scanner_name + "_combined.xlsx"
            baseline_file_path = os.path.join(self.results_base_directory,scanner_name,file_name)
            self.wb = openpyxl.load_workbook(baseline_file_path)
            self.ws = self.wb[coil_name]
            for col in self.ws[1]:
                baseline_headings.append(col.value)
            
            headings = results_dict["Combined"][date][acq_time]
            for heading in headings:
                if "Group" in heading:
                    headings_to_print.append(heading)
            #get combined element results for headngs of interest
            for heading in headings_to_print:
                slice_n = heading[-1]
                heading_ind = baseline_headings.index(heading)
                baseline_result = round(self.ws.cell(column=heading_ind+1, row=2).value,2)
                try:
                    results_to_print_c[slice_n].update({"measured "+heading[6:-8]:[results_dict["Combined"][date][acq_time][heading]], "baseline "+heading[6:-8]:[baseline_result]})
                except KeyError:
                    results_to_print_c[slice_n] = {"Element":["Combined"]}
                    results_to_print_c[slice_n].update({"measured "+heading[6:-8]:[results_dict["Combined"][date][acq_time][heading]], "baseline "+heading[6:-8]:[baseline_result]})        
            self.wb.close()
            
            baseline_headings = []
            file_name = "baseline_noise_av_" + scanner_name + "_delrec.xlsx"
            baseline_file_path = os.path.join(self.results_base_directory,scanner_name,file_name)
            self.wb = openpyxl.load_workbook(baseline_file_path)
            for slice_n in results_dict["DelRec"]:
                results_to_print_a[str(slice_n)] = {}
                headings_to_print = []
                ws_name = coil_name+"_slice_"+str(slice_n)
                self.ws = self.wb[ws_name]
                for col in self.ws[1]:
                    baseline_headings.append(col.value)
                headings = results_dict["DelRec"][slice_n][date][acq_time]
                for heading in headings:
                    if "Group" in heading:
                        headings_to_print.append(heading)
                
                for heading in headings_to_print:
                    
                    element_n = heading[heading.index("element_")+len("element_"):]
                    heading_ind = baseline_headings.index(heading)
                    baseline_result = round(self.ws.cell(column=heading_ind+1, row=2).value,2)
                    #get individual element results for headings of interest
                    try:
                        results_to_print_a[str(slice_n)]["Element"].append(int(element_n))
                        results_to_print_a[str(slice_n)]["measured SNR"].append(results_dict["DelRec"][slice_n][date][acq_time][heading])
                        results_to_print_a[str(slice_n)]["baseline SNR"].append(baseline_result)
                        results_to_print_a[str(slice_n)]["measured uniformity"].append("N/A")
                        results_to_print_a[str(slice_n)]["baseline uniformity"].append("N/A")
                    except KeyError:
                        results_to_print_a[str(slice_n)]["Element"] = [int(element_n)]
                        results_to_print_a[str(slice_n)]["measured SNR"] = [results_dict["DelRec"][slice_n][date][acq_time][heading]]
                        results_to_print_a[str(slice_n)]["baseline SNR"] = [baseline_result]
                        results_to_print_a[str(slice_n)]["measured uniformity"] = ["N/A"]
                        results_to_print_a[str(slice_n)]["baseline uniformity"] = ["N/A"]
                        
            self.wb.close()
            
            #folder to save png files in
            folder_name = os.path.join(self.png_archive,scanner_name,coil_name, date, "tables")
            if not os.path.exists(folder_name):
                os.makedirs(folder_name)
            for slice_no in results_to_print_a:
                table_a = pd.DataFrame(data=results_to_print_a[slice_no])
                table_a_sort = table_a.sort_values(by=["Element"])
                table_c = pd.DataFrame(data=results_to_print_c[slice_no])
                table_i = pd.concat([table_c, table_a_sort])
                table_i.reset_index(drop=True, inplace=True)
                png_file_name = "Slice_"+slice_no+".png"
                png_file_path = os.path.join(folder_name, png_file_name)
                dfi.export(table_i,png_file_path)
                
                
        def create_graphs(self,scanner_name,coil_name,use_group_av=False):
            """#if use group average is True one values only the group averages 
            will be printed, otherwise each individual result is plotted.
            create graph of current and historic data. Requires baseline 
            files to have been created"""
            baseline_headings = []
            headings_to_add = []
            baselines = {"SNR":{},"Uniformity":{}}
            #file name containing baseline data
            file_name = "baseline_noise_av_" + scanner_name + "_combined.xlsx"
            baseline_file_path = os.path.join(self.results_base_directory,scanner_name,file_name)
            self.wb = openpyxl.load_workbook(baseline_file_path)
            self.ws = self.wb[coil_name]
            for col in self.ws[1]:
                baseline_headings.append(col.value)
            for heading in baseline_headings:
                if "Group" in heading or "Uniformity" in heading:
                    headings_to_add.append(heading)
            
            for heading in headings_to_add:
                slice_n = heading[-1]
                if "SNR" in heading:
                    heading_ind = baseline_headings.index(heading)
                    baseline_result = self.ws.cell(column=heading_ind+1, row=2).value
                    baselines["SNR"][slice_n] = baseline_result
                if "uniformity" in heading:
                    heading_ind = baseline_headings.index(heading)
                    baseline_result = self.ws.cell(column=heading_ind+1, row=2).value
                    baselines["Uniformity"][slice_n] = baseline_result
            
            self.wb.close()
            
            
            
            if use_group_av == True:
                combined_file_name = "noise_av_" + scanner_name + "_group_combined.xlsx"
                delrec_file_name = "noise_av_" + scanner_name + "_group_delrec.xlsx"
            else:
                combined_file_name = "noise_av_" + scanner_name + "_combined.xlsx"
                delrec_file_name = "noise_av_" + scanner_name + "_delrec.xlsx"
            combined_file_path = os.path.join(self.results_base_directory,scanner_name,combined_file_name)
            delrec_file_path = os.path.join(self.results_base_directory,scanner_name,delrec_file_name)
            self.wb = openpyxl.load_workbook(combined_file_path)
            self.ws = self.wb[coil_name]
            combined_headings = []
            for col in self.ws[1]:
                combined_headings.append(col.value)
                        
            max_r = self.ws.max_row
            
            if use_group_av == True:
                for heading in combined_headings:
                    if "SNR" in heading or "uniformity" in heading:
                        slice_n = heading[-1] 
                        fig, ax = plt.subplots(nrows=1, ncols=1)
                        fig.set_tight_layout(True)
                        if "SNR" in heading:
                            baseline = baselines["SNR"][slice_n]
                            png_file_name = "Combined_SNR_Slice_"+slice_n+".png"
                            ax.title.set_text("Combined SNR")
                            ax.set_ylabel("SNR")
                            ax.set_xlabel("Date")
                        elif "uniformity" in heading:
                            baseline = baselines["Uniformity"][slice_n]
                            png_file_name = "Uniformity_Slice_"+slice_n+".png"
                            ax.title.set_text(" Uniformity")
                            ax.set_ylabel("Uniformity (%)")
                            ax.set_xlabel("Date")
                            
                        
                        heading_ind = combined_headings.index(heading)
                        
                        result_values = []
                        result_dates = []
                        for r in range(2,max_r+1):
                            result_values.append(self.ws.cell(row=r, column=heading_ind+1).value)
                            result_dates.append(self.ws.cell(row=r, column=1).value[-2:]+"/"+self.ws.cell(row=r, column=1).value[4:-2]+"/"+self.ws.cell(row=r, column=1).value[2:-4])
                        date = self.ws.cell(row=max_r, column=1).value
                        folder_name = os.path.join(self.png_archive,scanner_name,coil_name, date, "graphs")
                        if not os.path.exists(folder_name):
                            os.makedirs(folder_name)                    
                        png_file_path = os.path.join(folder_name, png_file_name)
                        x_values = np.arange(max_r-1)
                        ax.plot(x_values ,result_values, label='Results')
                        ax.hlines(y=baseline*0.9, xmin=x_values[0], xmax=x_values[-1], linewidth=2, color='r', linestyle="--", label='Baseline -10%')
                        ax.hlines(y=baseline*1.1, xmin=x_values[0], xmax=x_values[-1], linewidth=2, color='r', linestyle=":", label='Baseline +10%')
                        ax.set_xticks(x_values)
                        ax.set_xticklabels(result_dates, rotation=45)
                        ax.legend()
                        fig.savefig(png_file_path)
                        plt.close(fig)
            else:
                for heading in combined_headings:
                    if "Group" in heading:
                        pass
                    elif "SNR" in heading or "Uniformity" in heading:
                        slice_n = heading[-1] 
                        fig, ax = plt.subplots(nrows=1, ncols=1)
                        fig.set_tight_layout(True)
                        if "SNR" in heading:
                            baseline = baselines["SNR"][slice_n]
                            png_file_name = "Combined_SNR_Slice_"+slice_n+".png"
                            ax.title.set_text("Combined SNR")
                            ax.set_ylabel("SNR")
                            ax.set_xlabel("Date")
                        elif "Uniformity" in heading:
                            baseline = baselines["Uniformity"][slice_n]
                            png_file_name = "Uniformity_Slice_"+slice_n+".png"
                            ax.title.set_text(" Uniformity")
                            ax.set_ylabel("Uniformity (%)")
                            ax.set_xlabel("Date")
                            
                        
                        heading_ind = combined_headings.index(heading)
                        
                        result_values = []
                        result_dates = []
                        for r in range(2,max_r+1):
                            result_values.append(self.ws.cell(row=r, column=heading_ind+1).value)
                            result_dates.append(self.ws.cell(row=r, column=1).value[-2:]+"/"+self.ws.cell(row=r, column=1).value[4:-2]+"/"+self.ws.cell(row=r, column=1).value[2:-4])
                        date = self.ws.cell(row=max_r, column=1).value
                        folder_name = os.path.join(self.png_archive,scanner_name,coil_name, date, "graphs")
                        if not os.path.exists(folder_name):
                            os.makedirs(folder_name)                    
                        png_file_path = os.path.join(folder_name, png_file_name)
                        x_values = np.arange(max_r-1)
                        ax.plot(x_values ,result_values, label='Results')
                        ax.hlines(y=baseline*0.9, xmin=x_values[0], xmax=x_values[-1], linewidth=2, color='r', linestyle="--", label='Baseline -10%')
                        ax.hlines(y=baseline*1.1, xmin=x_values[0], xmax=x_values[-1], linewidth=2, color='r', linestyle=":", label='Baseline +10%')
                        ax.set_xticks(x_values)
                        ax.set_xticklabels(result_dates, rotation=45)
                        ax.legend()
                        fig.savefig(png_file_path)
                        plt.close(fig)
                    
            self.wb.close()
            
            
            
            baselines = {}
            baseline_file_name = "baseline_noise_av_" + scanner_name + "_delrec.xlsx"
            baseline_file_path = os.path.join(self.results_base_directory,scanner_name,baseline_file_name)
            self.wb = openpyxl.load_workbook(baseline_file_path)
            sheet_names = self.wb.sheetnames
            for sheet_name in sheet_names:
                if coil_name+"_slice_" in sheet_name:
                    baseline_headings = []
                    headings_to_add = []
                    slice_n = sheet_name[-1]
                    baselines[slice_n] = {}
                    self.ws = self.wb[sheet_name]
                    for col in self.ws[1]:
                        baseline_headings.append(col.value)
                    for heading in baseline_headings:
                        if "Group" in heading:
                            headings_to_add.append(heading)
                
                    for heading in headings_to_add:
                        element_n = heading[heading.index("element_")+len("element_"):]
                        heading_ind = baseline_headings.index(heading)
                        baseline_result = self.ws.cell(column=heading_ind+1, row=2).value
                        baselines[slice_n][element_n] = baseline_result
                        
            self.wb.close()
            
            
            
            
            self.wb = openpyxl.load_workbook(delrec_file_path)
            
            for slice_n in baselines:
                self.ws = self.wb[coil_name+"_slice_"+slice_n]
                delrec_headings = []
                for col in self.ws[1]:
                    delrec_headings.append(col.value)
            
                max_r = self.ws.max_row
                if use_group_av == True:
                    for heading in delrec_headings:
                        if "SNR" in heading:
                            element_n = heading[heading.index("element_")+len("element_"):]
                            fig, ax = plt.subplots(nrows=1, ncols=1)
                            fig.set_tight_layout(True)
                            baseline = baselines[slice_n][element_n]
                            png_file_name = "Element_"+element_n+"_SNR_Slice_"+slice_n+".png"
                            ax.title.set_text(" Element " + element_n + " SNR")
                            ax.set_ylabel("SNR")
                            ax.set_xlabel("Date")
                                
                            
                            heading_ind = delrec_headings.index(heading)
                            
                            result_values = []
                            result_dates = []
                            for r in range(2,max_r+1):
                                result_values.append(self.ws.cell(row=r, column=heading_ind+1).value)
                                result_dates.append(self.ws.cell(row=r, column=1).value[-2:]+"/"+self.ws.cell(row=r, column=1).value[4:-2]+"/"+self.ws.cell(row=r, column=1).value[2:-4])
                            date = self.ws.cell(row=max_r, column=1).value
                            png_file_path = os.path.join(self.png_archive,scanner_name,coil_name, date, "graphs", png_file_name)
                            x_values = np.arange(max_r-1)
                            ax.plot(x_values ,result_values, label='Results')
                            ax.hlines(y=baseline*0.90, xmin=x_values[0], xmax=x_values[-1], linewidth=2, color='r', linestyle="--", label='Baseline -10%')
                            ax.hlines(y=baseline*1.10, xmin=x_values[0], xmax=x_values[-1], linewidth=2, color='r', linestyle=":", label='Baseline +10%')
                            ax.set_xticks(x_values)
                            ax.set_xticklabels(result_dates, rotation=45)
                            ax.legend()
                            fig.savefig(png_file_path)
                            plt.close(fig)
                else:
                    for heading in delrec_headings:
                        if "Group" in heading:
                            pass
                        elif "SNR" in heading:
                            element_n = heading[heading.index("element_")+len("element_"):]
                            fig, ax = plt.subplots(nrows=1, ncols=1)
                            fig.set_tight_layout(True)
                            baseline = baselines[slice_n][element_n]
                            png_file_name = "Element_"+element_n+"_SNR_Slice_"+slice_n+".png"
                            ax.title.set_text(" Element " + element_n + " SNR")
                            ax.set_ylabel("SNR")
                            ax.set_xlabel("Date")
                                
                            
                            heading_ind = delrec_headings.index(heading)
                            
                            result_values = []
                            result_dates = []
                            for r in range(2,max_r+1):
                                result_values.append(self.ws.cell(row=r, column=heading_ind+1).value)
                                result_dates.append(self.ws.cell(row=r, column=1).value[-2:]+"/"+self.ws.cell(row=r, column=1).value[4:-2]+"/"+self.ws.cell(row=r, column=1).value[2:-4])
                            date = self.ws.cell(row=max_r, column=1).value
                            png_file_path = os.path.join(self.png_archive,scanner_name,coil_name, date, "graphs", png_file_name)
                            x_values = np.arange(max_r-1)
                            ax.plot(x_values ,result_values, label='Results')
                            ax.hlines(y=baseline*0.90, xmin=x_values[0], xmax=x_values[-1], linewidth=2, color='r', linestyle="--", label='Baseline -10%')
                            ax.hlines(y=baseline*1.10, xmin=x_values[0], xmax=x_values[-1], linewidth=2, color='r', linestyle=":", label='Baseline +10%')
                            ax.set_xticks(x_values)
                            ax.set_xticklabels(result_dates, rotation=45)
                            ax.legend()
                            fig.savefig(png_file_path)
                            plt.close(fig)
                        
                        
                        
                        
                        
                        
                    
                    
                    
                    
            self.wb.close()
                
                
                
                
                
                
                
                
                
            
                    
                    
            
            
            
    def __init__(self, master):
        """
        Parameters
        ----------
        master : tkinter root
        """
        master.withdraw()
        initialised = self.initialise_analysis(self.Quarterly_path)
        masks = self.initialise_masks(initialised.sorted_dcm_dict, initialised.lower_threshold, initialised.upper_threshold).mask_dict
        SNRs = self.calculate_results(initialised.sorted_dcm_dict, masks, initialised.n_elements)
        if initialised.excel_export == True:
            self.export_to_excel(SNRs, initialised.scanner_ID, initialised.coil_name)
        if initialised.figures == True:
            self.produce_figures(SNRs.sorted_SNR_results["noise_av"],initialised.scanner_name, initialised.coil_name)
        
        
        master.destroy()
        
        
main(root)

root.mainloop()

